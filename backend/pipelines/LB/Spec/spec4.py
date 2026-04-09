
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
OUTDIR = BASE / "lb_spec_outputs_v4"

def build_dataset_metadata():
    return pd.DataFrame([
        {
            "dataset_name": "LB",
            "label": "Laboratory Test Results",
            "class": "Findings",
            "structure": "One record per subject per lab test per final selected collection event/result record.",
            "domain_keys": "STUDYID, USUBJID, LBTESTCD, LBSEQ",
            "sort_order": "STUDYID, USUBJID, LBTESTCD, LBDTC, VISITNUM, LBSEQ",
            "standard": "SDTMIG v3.3-aligned implementation design",
            "implementation_note": "This spec is intended to drive SDTM LB generation, exception routing, row selection, and downstream define.xml metadata authoring.",
            "final_record_selection_rule": "Only rows passing Layer 1 QC and deterministic mapping rules enter candidate LB build. Rows with unresolved contradictions go to exception dataset, not final LB."
        }
    ])

def build_main_spec():
    rows = [
        {
            "spec_seq": 1, "target_domain": "LB", "target_variable": "STUDYID",
            "target_label": "Study Identifier", "target_type": "Char", "target_length": 20,
            "core": "Req", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "PROTOCOL_NO",
            "source_label": "Protocol Number", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "",
            "rule": "Map STUDYID = strip(PROTOCOL_NO). Preserve the collected protocol identifier as character.",
            "primary_qc_checks": "PROTOCOL_NO must not be missing and must be constant across the study.",
            "traceability_note": "Direct trace to source field PROTOCOL_NO.",
            "programming_notes": "Do not cast to numeric.",
            "definexml_origin_comment": "Collected on CRF/vendor feed.",
            "exception_condition": "Missing or inconsistent protocol -> exception",
            "final_condition": "Non-missing and study-consistent", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 2, "target_domain": "LB", "target_variable": "DOMAIN",
            "target_label": "Domain Abbreviation", "target_type": "Char", "target_length": 2,
            "core": "Req", "origin": "Assigned", "mapping_class": "Constant",
            "source_form_or_module": "Not collected / Assigned", "source_variable": "Not Applicable",
            "source_label": "Not Applicable", "source_type_hint": "N/A", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "", "reconciliation_source": "",
            "rule": 'Set DOMAIN = "LB" for all records.',
            "primary_qc_checks": 'DOMAIN must equal "LB" for every output row.',
            "traceability_note": "Assigned constant for LB domain build.",
            "programming_notes": "No source field is used.",
            "definexml_origin_comment": "Assigned constant.",
            "exception_condition": 'If DOMAIN is not "LB" -> exception',
            "final_condition": 'Always "LB"', "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 3, "target_domain": "LB", "target_variable": "USUBJID",
            "target_label": "Unique Subject Identifier", "target_type": "Char", "target_length": 40,
            "core": "Req", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "PROTOCOL_NO + SITE_NO + SUBJECT_NO",
            "source_label": "Protocol + Site + Subject", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "SUBJECT_KEY",
            "rule": 'If SUBJECT_KEY is present and sponsor-compliant, preserve it. Otherwise derive USUBJID = PROTOCOL_NO || "-" || SITE_NO || "-" || SUBJECT_NO, preserving site and subject identifiers as character strings.',
            "primary_qc_checks": "USUBJID must not be missing. Final LB output must not contain duplicate STUDYID/USUBJID/LBTESTCD/LBSEQ combinations.",
            "traceability_note": "Traceable to protocol/site/subject identifying fields in raw LB.",
            "programming_notes": "Do not coerce SITE_NO or SUBJECT_NO to numeric if left-padding may be lost.",
            "definexml_origin_comment": "Derived from collected identifiers.",
            "exception_condition": "Missing or duplicate USUBJID -> exception",
            "final_condition": "Unique and non-missing", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 4, "target_domain": "LB", "target_variable": "LBSEQ",
            "target_label": "Sequence Number", "target_type": "Num", "target_length": 8,
            "core": "Req", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Assigned", "source_variable": "Sorted final LB record order",
            "source_label": "Operational sequence", "source_type_hint": "numeric", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "",
            "rule": "After final LB record selection and sorting by STUDYID, USUBJID, LBTESTCD, LBDTC, VISITNUM, assign LBSEQ as a sequential integer starting at 1 within each USUBJID.",
            "primary_qc_checks": "LBSEQ must be non-missing, numeric, and unique within USUBJID.",
            "traceability_note": "Derived operational sequence for final SDTM LB.",
            "programming_notes": "Assign after de-duplication, row selection, and exception handling.",
            "definexml_origin_comment": "Derived by algorithm after final row selection.",
            "exception_condition": "Duplicate or missing LBSEQ -> exception",
            "final_condition": "Unique within subject", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 5, "target_domain": "LB", "target_variable": "LBTESTCD",
            "target_label": "Lab Test or Examination Short Name", "target_type": "Char", "target_length": 8,
            "core": "Req", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "TEST_CODE_RAW",
            "source_label": "Raw Lab Test Code", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "LBTESTCD codelist / lb_test_map_v4.csv", "reconciliation_source": "TEST_NAME_RAW",
            "rule": "Map TEST_CODE_RAW to approved LBTESTCD using lb_test_map_v4.csv. If raw code is not directly mapped but TEST_NAME_RAW matches an approved synonym for the same test, map using the approved combination. Unmapped or contradictory code/name pairs go to exception.",
            "primary_qc_checks": "LBTESTCD must not be missing in final. TEST_CODE_RAW and TEST_NAME_RAW must not contradict the approved test map.",
            "traceability_note": "Traceable through the approved test-mapping sheet.",
            "programming_notes": "Use deterministic mapping only. Do not guess unmapped tests.",
            "definexml_origin_comment": "Assigned from sponsor-controlled test mapping.",
            "exception_condition": "Unmapped or contradictory test code/name -> exception",
            "final_condition": "Mapped to approved LBTESTCD", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 6, "target_domain": "LB", "target_variable": "LBTEST",
            "target_label": "Lab Test or Examination Name", "target_type": "Char", "target_length": 40,
            "core": "Req", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "TEST_NAME_RAW",
            "source_label": "Raw Lab Test Name", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "LBTEST codelist / lb_test_map_v4.csv", "reconciliation_source": "TEST_CODE_RAW",
            "rule": "Map TEST_NAME_RAW to approved LBTEST using lb_test_map_v4.csv. When both TEST_CODE_RAW and TEST_NAME_RAW are present, they must resolve to the same standard test. Otherwise, route to exception.",
            "primary_qc_checks": "LBTEST must not be missing in final and must align one-to-one with LBTESTCD.",
            "traceability_note": "Traceable through the approved test-mapping sheet.",
            "programming_notes": "Normalize only via approved synonyms and code/name pairings.",
            "definexml_origin_comment": "Assigned from sponsor-controlled test mapping.",
            "exception_condition": "Unmapped or contradictory raw test name -> exception",
            "final_condition": "Mapped to approved LBTEST", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 7, "target_domain": "LB", "target_variable": "LBCAT",
            "target_label": "Category for Lab Test", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "TEST_PANEL_RAW",
            "source_label": "Raw Test Panel", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_test_map_v4.csv", "reconciliation_source": "TEST_CODE_RAW",
            "rule": "Map TEST_PANEL_RAW to approved LBCAT using lb_test_map_v4.csv. If TEST_PANEL_RAW is missing but the mapped test has a single approved category, populate from the mapping sheet. Contradictory panel/test combinations go to exception.",
            "primary_qc_checks": "If populated, LBCAT should align with the approved category for the mapped LBTESTCD.",
            "traceability_note": "Traceable through raw panel and approved test map.",
            "programming_notes": "Do not silently override contradictory panel values.",
            "definexml_origin_comment": "Assigned from sponsor-controlled panel mapping.",
            "exception_condition": "Contradictory panel/test combination -> exception",
            "final_condition": "Missing or mapped approved category", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 8, "target_domain": "LB", "target_variable": "LBSPEC",
            "target_label": "Specimen Material Type", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "SPECIMEN_RAW",
            "source_label": "Raw Specimen", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_specimen_map_v4.csv", "reconciliation_source": "TEST_CODE_RAW",
            "rule": "Map SPECIMEN_RAW to approved LBSPEC values using lb_specimen_map_v4.csv. Approved values in this demo include SERUM, PLASMA, BLOOD, WHOLE BLOOD, and URINE. If the specimen is incompatible with the mapped test, route to exception.",
            "primary_qc_checks": "LBSPEC should be present when collected and should be compatible with the mapped lab test.",
            "traceability_note": "Traceable to raw specimen field with controlled-value normalization.",
            "programming_notes": "Do not infer specimen when missing unless sponsor explicitly approves a default by test.",
            "definexml_origin_comment": "Collected and standardized to controlled terminology.",
            "exception_condition": "Invalid or incompatible specimen -> exception",
            "final_condition": "Missing or approved specimen value", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 9, "target_domain": "LB", "target_variable": "LBLOINC",
            "target_label": "Logical Observation Identifiers Names and Codes", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Constant",
            "source_form_or_module": "Not explicitly collected or mapped in this demo", "source_variable": "Not collected",
            "source_label": "Not collected", "source_type_hint": "N/A", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "LOINC", "reconciliation_source": "",
            "rule": "Leave LBLOINC blank in this demo implementation because no approved LOINC mapping table is provided. Populate only if sponsor later supplies a deterministic LOINC mapping by LBTESTCD/LBTEST.",
            "primary_qc_checks": "Blank is acceptable in this demo design.",
            "traceability_note": "No source or approved mapping available in current design.",
            "programming_notes": "Keep metadata row for completeness and define.xml readiness.",
            "definexml_origin_comment": "No approved LOINC mapping in this implementation; variable not populated.",
            "exception_condition": "Not applicable",
            "final_condition": "Blank", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 10, "target_domain": "LB", "target_variable": "LBORRES",
            "target_label": "Result or Finding in Original Units", "target_type": "Char", "target_length": 20,
            "core": "Req", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "RESULT_RAW",
            "source_label": "Raw Result", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "RESULT_NUM_RAW + RESULT_CHAR_RAW + RESULT_QUAL_RAW",
            "rule": "Map LBORRES = RESULT_RAW preserving the collected character representation exactly after trim. This includes numeric-looking values, categorical values such as NEGATIVE/POSITIVE/TRACE, and comparator-style results such as <0.3.",
            "primary_qc_checks": "LBORRES must not be missing for performed records. If LBSTAT='NOT DONE', result-related variables must not populate in final LB.",
            "traceability_note": "Direct trace to collected raw result field.",
            "programming_notes": "Preserve the original character result for auditability; do not standardize LBORRES.",
            "definexml_origin_comment": "Collected result value.",
            "exception_condition": "Missing LBORRES on a performed record -> exception",
            "final_condition": "Non-missing for valid performed record", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 11, "target_domain": "LB", "target_variable": "LBORRESU",
            "target_label": "Original Units", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "ORIG_UNIT_RAW",
            "source_label": "Original Result Unit", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_unit_map_v4.csv", "reconciliation_source": "TEST_CODE_RAW",
            "rule": "Normalize ORIG_UNIT_RAW to approved LBORRESU text using lb_unit_map_v4.csv without converting the physical value. For categorical urinalysis-style results without meaningful units, LBORRESU may remain blank.",
            "primary_qc_checks": "For numeric tests, LBORRESU should be present and approved for the mapped test unless the row is routed to exception.",
            "traceability_note": "Traceable through unit normalization mapping sheet.",
            "programming_notes": "Normalize text only. Do not convert values in LBORRESU.",
            "definexml_origin_comment": "Collected unit normalized to controlled terminology.",
            "exception_condition": "Unsupported or missing unit for numeric test -> exception",
            "final_condition": "Missing or approved normalized original unit", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 12, "target_domain": "LB", "target_variable": "LBORNRLO",
            "target_label": "Reference Range Lower Limit in Orig Unit", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "REF_LOW_RAW",
            "source_label": "Raw Reference Low", "source_type_hint": "numeric-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "REF_UNIT_RAW + ORIG_UNIT_RAW",
            "rule": "If REF_LOW_RAW is numeric and the range pertains to the original result context, map directly to LBORNRLO after numeric interpretation. If the raw range is non-numeric, invalid, or clearly incompatible with the original result context, leave blank and route to exception/QC.",
            "primary_qc_checks": "LBORNRLO should be numeric when populated. REF_LOW_RAW >= REF_HIGH_RAW goes to exception.",
            "traceability_note": "Direct trace to raw original-unit reference range lower limit.",
            "programming_notes": "Do not force populate when range context is ambiguous.",
            "definexml_origin_comment": "Collected reference range lower limit in original units when usable.",
            "exception_condition": "Invalid or incompatible original reference range -> exception/QC",
            "final_condition": "Blank or numeric lower limit", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 13, "target_domain": "LB", "target_variable": "LBORNRHI",
            "target_label": "Reference Range Upper Limit in Orig Unit", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "REF_HIGH_RAW",
            "source_label": "Raw Reference High", "source_type_hint": "numeric-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "REF_UNIT_RAW + ORIG_UNIT_RAW",
            "rule": "If REF_HIGH_RAW is numeric and the range pertains to the original result context, map directly to LBORNRHI after numeric interpretation. If the raw range is non-numeric, invalid, or clearly incompatible with the original result context, leave blank and route to exception/QC.",
            "primary_qc_checks": "LBORNRHI should be numeric when populated. REF_LOW_RAW >= REF_HIGH_RAW goes to exception.",
            "traceability_note": "Direct trace to raw original-unit reference range upper limit.",
            "programming_notes": "Do not force populate when range context is ambiguous.",
            "definexml_origin_comment": "Collected reference range upper limit in original units when usable.",
            "exception_condition": "Invalid or incompatible original reference range -> exception/QC",
            "final_condition": "Blank or numeric upper limit", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 14, "target_domain": "LB", "target_variable": "LBSTRESC",
            "target_label": "Character Result/Finding in Std Format", "target_type": "Char", "target_length": 20,
            "core": "Exp", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "RESULT_RAW + RESULT_NUM_RAW + RESULT_CHAR_RAW + RESULT_QUAL_RAW",
            "source_label": "Raw result components", "source_type_hint": "mixed", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "TEST_CODE_RAW + ORIG_UNIT_RAW",
            "rule": "Derive LBSTRESC as the standardized character representation of the result. For numeric tests with approved standardization, LBSTRESC is the character form of the standardized numeric result. For categorical tests, map approved categorical results to the standardized character result. Preserve comparator meaning in LBORRES; do not include comparator symbols in LBSTRESC unless sponsor explicitly requires it.",
            "primary_qc_checks": "LBSTRESC should align with LBSTRESN when numeric standardization succeeds.",
            "traceability_note": "Traceable through result standardization rules and raw result components.",
            "programming_notes": "Do not derive LBSTRESC when the raw result is ambiguous or routed to exception.",
            "definexml_origin_comment": "Derived by deterministic standardization algorithm.",
            "exception_condition": "Standardization cannot be performed deterministically -> exception",
            "final_condition": "Standardization succeeds", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Comparator handling may vary by sponsor; this demo keeps comparator only in LBORRES."
        },
        {
            "spec_seq": 15, "target_domain": "LB", "target_variable": "LBSTRESN",
            "target_label": "Numeric Result/Finding in Standard Units", "target_type": "Num", "target_length": 8,
            "core": "Exp", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "RESULT_NUM_RAW + RESULT_RAW + ORIG_UNIT_RAW + TEST_CODE_RAW",
            "source_label": "Raw numeric result + unit + test", "source_type_hint": "mixed", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_unit_map_v4.csv", "reconciliation_source": "REF_UNIT_RAW",
            "rule": "Derive LBSTRESN only when the mapped test is numeric and the raw result can be interpreted numerically in an approved unit. If RESULT_NUM_RAW is present and consistent with RESULT_RAW, use RESULT_NUM_RAW as the starting numeric value. If the original and standard units are the same, carry the numeric value forward. If a deterministic unit conversion rule exists in lb_unit_map_v4.csv, convert accordingly. If the result is non-numeric, comparator-only, categorical, or has conflicting units not approved for conversion, route to exception.",
            "primary_qc_checks": "LBSTRESN must be numeric when populated. Unsupported units, missing units for numeric tests, or ambiguous numeric interpretations go to exception.",
            "traceability_note": "Traceable through approved unit conversion rules and raw result components.",
            "programming_notes": "Do not derive LBSTRESN from categorical results such as NEGATIVE/POSITIVE/TRACE.",
            "definexml_origin_comment": "Derived by deterministic standardization algorithm.",
            "exception_condition": "Unsupported unit, ambiguous value, or non-numeric result for numeric test -> exception",
            "final_condition": "Numeric derivation succeeds deterministically", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "This demo supports only explicitly approved conversions."
        },
        {
            "spec_seq": 16, "target_domain": "LB", "target_variable": "LBSTRESU",
            "target_label": "Standard Units", "target_type": "Char", "target_length": 20,
            "core": "Exp", "origin": "Assigned", "mapping_class": "Derived",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "TEST_CODE_RAW + ORIG_UNIT_RAW",
            "source_label": "Raw test + raw unit", "source_type_hint": "mixed", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_unit_map_v4.csv", "reconciliation_source": "",
            "rule": "Assign LBSTRESU by mapped test according to approved standard units from lb_unit_map_v4.csv. If the source unit is already the standard unit, carry it forward after normalization. If the row is categorical or does not support deterministic standardization, LBSTRESU may remain blank.",
            "primary_qc_checks": "When LBSTRESN is populated, LBSTRESU should also be populated and consistent with the mapped standard test/unit rules.",
            "traceability_note": "Traceable through standard-unit assignment rules by test.",
            "programming_notes": "Do not populate LBSTRESU when standardization is not supported.",
            "definexml_origin_comment": "Assigned from approved standard-unit metadata as part of derivation.",
            "exception_condition": "Unsupported standard unit assignment -> exception",
            "final_condition": "Blank or approved standard unit", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 17, "target_domain": "LB", "target_variable": "LBSTNRLO",
            "target_label": "Reference Range Lower Limit-Std Units", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Derived", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "REF_LOW_RAW",
            "source_label": "Raw Reference Low", "source_type_hint": "numeric-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "REF_UNIT_RAW + ORIG_UNIT_RAW",
            "rule": "If REF_LOW_RAW is numeric and its unit context is consistent with the unit used for result standardization, map or convert to LBSTNRLO. If the reference range units are incompatible with the result standardization or the range is missing/invalid, leave LBSTNRLO blank and route the row to exception/QC according to sponsor policy.",
            "primary_qc_checks": "LBSTNRLO should be numeric when populated. Rows with REF_LOW_RAW >= REF_HIGH_RAW go to exception.",
            "traceability_note": "Traceable to raw reference range fields.",
            "programming_notes": "Do not force range conversion when unit context is ambiguous.",
            "definexml_origin_comment": "Derived/conformed from collected range in standard unit context.",
            "exception_condition": "Invalid or incompatible range context -> exception/QC",
            "final_condition": "Blank or numeric lower bound", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Reference range conversion policy may vary by sponsor."
        },
        {
            "spec_seq": 18, "target_domain": "LB", "target_variable": "LBSTNRHI",
            "target_label": "Reference Range Upper Limit-Std Units", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Derived", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "REF_HIGH_RAW",
            "source_label": "Raw Reference High", "source_type_hint": "numeric-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "REF_UNIT_RAW + ORIG_UNIT_RAW",
            "rule": "If REF_HIGH_RAW is numeric and its unit context is consistent with the unit used for result standardization, map or convert to LBSTNRHI. If the reference range units are incompatible with the result standardization or the range is missing/invalid, leave LBSTNRHI blank and route the row to exception/QC according to sponsor policy.",
            "primary_qc_checks": "LBSTNRHI should be numeric when populated. Rows with REF_LOW_RAW >= REF_HIGH_RAW go to exception.",
            "traceability_note": "Traceable to raw reference range fields.",
            "programming_notes": "Do not force range conversion when unit context is ambiguous.",
            "definexml_origin_comment": "Derived/conformed from collected range in standard unit context.",
            "exception_condition": "Invalid or incompatible range context -> exception/QC",
            "final_condition": "Blank or numeric upper bound", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Reference range conversion policy may vary by sponsor."
        },
        {
            "spec_seq": 19, "target_domain": "LB", "target_variable": "LBNRIND",
            "target_label": "Reference Range Indicator", "target_type": "Char", "target_length": 8,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "ABN_FLAG_RAW",
            "source_label": "Raw Abnormal Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "NORMAL | LOW | HIGH", "reconciliation_source": "RESULT_NUM_RAW + REF_LOW_RAW + REF_HIGH_RAW + ORIG_UNIT_RAW + REF_UNIT_RAW",
            "rule": "Map ABN_FLAG_RAW to approved LBNRIND values where possible. In this demo, N/NORMAL maps to NORMAL, L/LOW maps to LOW, and H/HIGH maps to HIGH. If ABN_FLAG_RAW is missing but numeric comparison to range is valid and deterministic, derive LBNRIND from the standardized comparison. If result/range/unit context is ambiguous, route to exception rather than derive.",
            "primary_qc_checks": "If numeric comparison is possible, LBNRIND should not contradict the result and reference range.",
            "traceability_note": "Traceable to raw abnormal flag and range comparison logic.",
            "programming_notes": "Prefer direct mapped abnormal flag when valid; only derive when comparison context is clean.",
            "definexml_origin_comment": "Collected flag with possible deterministic derivation when permitted.",
            "exception_condition": "Contradictory abnormal flag or ambiguous comparison -> exception",
            "final_condition": "Blank or approved range indicator", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 20, "target_domain": "LB", "target_variable": "LBDTC",
            "target_label": "Date/Time of Specimen Collection", "target_type": "Char", "target_length": 20,
            "core": "Req", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "COLL_DTM_RAW",
            "source_label": "Collection DateTime Raw", "source_type_hint": "datetime-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "ISO 8601", "reconciliation_source": "COLL_DATE_RAW + COLL_TIME_RAW",
            "rule": "If COLL_DTM_RAW is a valid ISO 8601 datetime, map directly to LBDTC. If COLL_DTM_RAW is non-ISO but valid and deterministic, standardize it to ISO 8601. If COLL_DTM_RAW is missing but COLL_DATE_RAW and COLL_TIME_RAW are both valid and consistent, derive LBDTC = COLL_DATE_RAW || 'T' || COLL_TIME_RAW after ISO standardization. If the datetime conflicts with the separate date/time fields, route to exception.",
            "primary_qc_checks": "LBDTC must not be missing for final performed records. Conflicting collection date/time fields go to exception.",
            "traceability_note": "Traceable to collection datetime and supporting date/time fields.",
            "programming_notes": "Preserve collected granularity when only date is available and sponsor-accepted; otherwise use full datetime when valid.",
            "definexml_origin_comment": "Collected date/time standardized to ISO 8601.",
            "exception_condition": "Missing, invalid, or contradictory collection datetime -> exception",
            "final_condition": "Valid ISO 8601 date/time or date per sponsor rule", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 21, "target_domain": "LB", "target_variable": "VISIT",
            "target_label": "Visit Name", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "VISIT_RAW",
            "source_label": "Visit Name Raw", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_visit_map_v4.csv", "reconciliation_source": "VISITNUM_RAW + UNSCHED_RAW",
            "rule": "Map VISIT = VISIT_RAW after trim. Preserve the sponsor-defined visit label. VISIT_RAW, VISITNUM_RAW, and UNSCHED_RAW must be consistent with lb_visit_map_v4.csv and unscheduled visit logic. Contradictions go to exception.",
            "primary_qc_checks": "VISIT and VISITNUM must align with the approved visit map.",
            "traceability_note": "Direct trace to raw visit field.",
            "programming_notes": "Do not silently recode a contradictory visit.",
            "definexml_origin_comment": "Collected visit label.",
            "exception_condition": "Visit-map contradiction -> exception",
            "final_condition": "Missing or approved visit label", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 22, "target_domain": "LB", "target_variable": "VISITNUM",
            "target_label": "Visit Number", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "VISITNUM_RAW",
            "source_label": "Visit Number Raw", "source_type_hint": "numeric-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "lb_visit_map_v4.csv", "reconciliation_source": "VISIT_RAW",
            "rule": "Map VISITNUM = VISITNUM_RAW after numeric normalization where the value is numeric-like. Preserve sponsor visit numbering. If VISIT_RAW and VISITNUM_RAW are inconsistent with lb_visit_map_v4.csv, route the row to exception rather than manufacture a corrected VISITNUM.",
            "primary_qc_checks": "VISITNUM should be numeric-like and aligned to the approved visit map.",
            "traceability_note": "Direct trace to raw visit-number field.",
            "programming_notes": "Do not derive VISITNUM from VISIT_RAW unless sponsor explicitly approves it.",
            "definexml_origin_comment": "Collected visit number.",
            "exception_condition": "Invalid or contradictory visit number -> exception",
            "final_condition": "Missing or approved numeric visit number", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 23, "target_domain": "LB", "target_variable": "LBFAST",
            "target_label": "Fasting Status", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "FASTING_RAW",
            "source_label": "Raw Fasting Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Y | N | U", "reconciliation_source": "",
            "rule": "Map FASTING_RAW to LBFAST using approved values Y and N in this demo. If sponsor later collects unknown fasting status, U may also be used in accordance with standard conventions. If FASTING_RAW is missing, leave LBFAST blank. Unsupported values go to exception.",
            "primary_qc_checks": "If populated, LBFAST must equal Y, N, or U.",
            "traceability_note": "Direct trace to raw fasting flag.",
            "programming_notes": "Current raw design supports Y/N only; U is reserved for future sponsor-approved use if collected.",
            "definexml_origin_comment": "Collected fasting status normalized to controlled terminology.",
            "exception_condition": "Unsupported fasting value -> exception",
            "final_condition": "Blank or Y/N/U", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 24, "target_domain": "LB", "target_variable": "LBSTAT",
            "target_label": "Completion Status", "target_type": "Char", "target_length": 12,
            "core": "Perm", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "NOT_DONE_RAW",
            "source_label": "Not Done Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "NOT DONE", "reconciliation_source": "ND_REASON_RAW",
            "rule": "If NOT_DONE_RAW = 'Y', set LBSTAT = 'NOT DONE'. Otherwise leave LBSTAT blank. Records marked not done should generally not populate result variables in final LB.",
            "primary_qc_checks": "If LBSTAT='NOT DONE', result-related fields should not populate in final LB.",
            "traceability_note": "Traceable to raw not-done flag.",
            "programming_notes": "Keep NOT DONE rows in final LB only when sponsor policy requires them; otherwise route to non-output exception handling.",
            "definexml_origin_comment": "Derived from collected NOT_DONE indicator.",
            "exception_condition": "NOT_DONE contradiction with populated result fields -> exception",
            "final_condition": "NOT DONE or blank", "review_status": "Final", "confidence": "High", "ambiguity_note": "Submission policy for not-done rows should be confirmed study-by-study."
        },
        {
            "spec_seq": 25, "target_domain": "LB", "target_variable": "LBREASND",
            "target_label": "Reason Not Done", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Laboratory Results CRF / Lab Vendor Feed", "source_variable": "ND_REASON_RAW",
            "source_label": "Not Done Reason", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "reconciliation_source": "NOT_DONE_RAW",
            "rule": "If NOT_DONE_RAW = 'Y', map LBREASND = ND_REASON_RAW after trim. If NOT_DONE_RAW is not 'Y', LBREASND should remain blank in final LB.",
            "primary_qc_checks": "LBREASND should only populate when LBSTAT='NOT DONE' or NOT_DONE_RAW='Y'.",
            "traceability_note": "Direct trace to raw not-done reason.",
            "programming_notes": "Do not populate LBREASND for performed rows.",
            "definexml_origin_comment": "Collected not-done reason.",
            "exception_condition": "Reason present without NOT DONE context -> exception",
            "final_condition": "Blank or trimmed reason", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 26, "target_domain": "LB", "target_variable": "LBMETHOD",
            "target_label": "Method of Test or Examination", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Constant",
            "source_form_or_module": "Not explicitly collected in this raw design", "source_variable": "Not collected",
            "source_label": "Not collected", "source_type_hint": "N/A", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "", "reconciliation_source": "",
            "rule": "Leave LBMETHOD blank in this demo because no explicit laboratory method field is collected. Populate only if sponsor later provides a deterministic source.",
            "primary_qc_checks": "Blank is acceptable in this demo design.",
            "traceability_note": "No source field available in current raw design.",
            "programming_notes": "Keep metadata row for completeness and define.xml readiness.",
            "definexml_origin_comment": "No source; variable intentionally left blank in this implementation.",
            "exception_condition": "Not applicable",
            "final_condition": "Blank", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
    ]
    return pd.DataFrame(rows)

def build_test_map():
    return pd.DataFrame([
        ["HGB", "HEMOGLOBIN", "Numeric hematology result", "HEMATOLOGY", "BLOOD|WHOLE BLOOD", "NUMERIC"],
        ["WBC", "WHITE BLOOD CELL COUNT", "Numeric hematology result", "HEMATOLOGY", "BLOOD|WHOLE BLOOD", "NUMERIC"],
        ["PLT", "PLATELET COUNT", "Numeric hematology result", "HEMATOLOGY", "BLOOD|WHOLE BLOOD", "NUMERIC"],
        ["PLAT", "PLATELET COUNT", "Raw synonym of PLT", "HEMATOLOGY", "BLOOD|WHOLE BLOOD", "NUMERIC"],
        ["ALT", "ALANINE AMINOTRANSFERASE", "Approved synonym pair", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["ALT", "ALT (SGPT)", "Approved synonym pair", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["AST", "ASPARTATE AMINOTRANSFERASE", "Approved synonym pair", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["AST", "AST (SGOT)", "Approved synonym pair", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["BILI", "BILIRUBIN TOTAL", "Numeric chemistry result", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["CREAT", "CREATININE", "Numeric chemistry result", "CHEMISTRY", "SERUM|PLASMA|URINE", "NUMERIC"],
        ["GLUC", "GLUCOSE", "Numeric chemistry result", "CHEMISTRY", "SERUM|PLASMA", "NUMERIC"],
        ["PT", "PROTHROMBIN TIME", "Numeric coagulation result", "COAGULATION", "PLASMA", "NUMERIC"],
        ["INR", "INR", "Numeric coagulation result", "COAGULATION", "PLASMA", "NUMERIC"],
        ["HCG", "HCG", "Numeric pregnancy result", "PREGNANCY", "SERUM|URINE", "NUMERIC"],
        ["UPROT", "URINE PROTEIN", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
        ["UPROT", "PROTEIN", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
        ["UGLUC", "URINE GLUCOSE", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
        ["UGLUC", "GLUCOSE", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
        ["KET", "KETONES", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
        ["BLOODU", "BLOOD", "Categorical urinalysis result", "URINALYSIS", "URINE", "CATEGORICAL"],
    ], columns=["raw_test_code", "raw_test_name", "mapping_note", "lbcat", "allowed_specimen", "result_type"])

def build_unit_map():
    return pd.DataFrame([
        ["HGB", "G/DL", "G/DL", "G/DL", "no conversion", "Yes"],
        ["WBC", "10^9/L", "10^9/L", "10^9/L", "no conversion", "Yes"],
        ["WBC", "CELLS/UL", "CELLS/UL", "CELLS/UL", "no conversion", "Yes"],
        ["PLT", "10^9/L", "10^9/L", "10^9/L", "no conversion", "Yes"],
        ["PLAT", "10^9/L", "10^9/L", "10^9/L", "no conversion", "Yes"],
        ["ALT", "U/L", "U/L", "U/L", "no conversion", "Yes"],
        ["AST", "U/L", "U/L", "U/L", "no conversion", "Yes"],
        ["BILI", "MG/DL", "MG/DL", "MG/DL", "no conversion", "Yes"],
        ["BILI", "UMOL/L", "UMOL/L", "MG/DL", "not implemented in demo", "No"],
        ["CREAT", "MG/DL", "MG/DL", "MG/DL", "no conversion", "Yes"],
        ["CREAT", "UMOL/L", "UMOL/L", "MG/DL", "not implemented in demo", "No"],
        ["GLUC", "MG/DL", "MG/DL", "MG/DL", "no conversion", "Yes"],
        ["GLUC", "MMOL/L", "MMOL/L", "MG/DL", "not implemented in demo", "No"],
        ["PT", "SEC", "SEC", "SEC", "no conversion", "Yes"],
        ["INR", "RATIO", "RATIO", "RATIO", "no conversion", "Yes"],
        ["HCG", "MIU/ML", "MIU/ML", "MIU/ML", "no conversion", "Yes"],
        ["HCG", "IU/L", "IU/L", "MIU/ML", "not implemented in demo", "No"],
    ], columns=["lbtestcd", "raw_unit", "lborresu_normalized", "lbstresu_standard", "conversion_rule", "implemented_in_demo"])

def build_visit_map():
    return pd.DataFrame([
        ["SCREENING", 10, "N", "Approved sponsor convention"],
        ["BASELINE", 10, "N", "Approved sponsor convention; same VISITNUM as SCREENING in this demo"],
        ["DAY 1", 20, "N", ""],
        ["WEEK 1", 20, "N", ""],
        ["WEEK 2", 30, "N", ""],
        ["WEEK 4", 40, "N", ""],
        ["UNSCHEDULED", 99, "Y", ""],
    ], columns=["visit_name", "visitnum", "unsched_expected", "note"])

def build_specimen_map():
    return pd.DataFrame([
        ["SERUM", "SERUM"],
        ["PLASMA", "PLASMA"],
        ["BLOOD", "BLOOD"],
        ["WHOLE BLOOD", "WHOLE BLOOD"],
        ["URINE", "URINE"],
    ], columns=["raw_specimen", "lbspec"])

def build_codelists():
    return pd.DataFrame([
        ["LBTESTCD", "Code list", "Controlled per lb_test_map_v4.csv"],
        ["LBTEST", "Code list", "Controlled per lb_test_map_v4.csv"],
        ["LBCAT", "Code list", "HEMATOLOGY | CHEMISTRY | COAGULATION | URINALYSIS | PREGNANCY"],
        ["LBSPEC", "Code list", "SERUM | PLASMA | BLOOD | WHOLE BLOOD | URINE"],
        ["LBFAST", "Code list", "Y | N | U"],
        ["LBSTAT", "Code list", "NOT DONE"],
        ["LBNRIND", "Code list", "NORMAL | LOW | HIGH"],
        ["VISIT", "Sponsor codelist", "Controlled per lb_visit_map_v4.csv"],
        ["LBLOINC", "External codelist", "LOINC if sponsor later provides approved mapping"],
    ], columns=["variable", "codelist_type", "permitted_values_or_reference"])

def build_value_level_metadata():
    return pd.DataFrame([
        ["LBSTRESN", "LBTESTCD in numeric-test set AND deterministic standard unit conversion supported", "Num", "Standardized numeric value only for numeric tests with implemented conversion or same-unit carry-forward", "lb_test_map_v4.csv + lb_unit_map_v4.csv"],
        ["LBSTRESU", "LBTESTCD in numeric-test set AND deterministic standard unit conversion supported", "Char", "Populate approved standard unit when LBSTRESN is derived", "lb_unit_map_v4.csv"],
        ["LBSTNRLO", "Numeric reference range available and compatible with standardization context", "Num", "Populate only when reference range is numeric and unit context is compatible", "Raw range fields + lb_unit_map_v4.csv"],
        ["LBSTNRHI", "Numeric reference range available and compatible with standardization context", "Num", "Populate only when reference range is numeric and unit context is compatible", "Raw range fields + lb_unit_map_v4.csv"],
        ["LBNRIND", "Numeric comparison context clean OR valid collected abnormal flag available", "Char", "Map collected flag or derive only when range comparison is deterministic", "ABN_FLAG_RAW + numeric/range logic"],
        ["LBORRESU", "Categorical urinalysis-style tests", "Char", "May remain blank for categorical results without meaningful units", "lb_test_map_v4.csv"],
        ["LBLOINC", "Sponsor later provides approved LBTESTCD/LBTEST to LOINC mapping", "Char", "Currently blank in this demo; metadata row retained for define.xml completeness", "Future mapping table if approved"],
    ], columns=["variable", "where_clause", "datatype", "comment", "source_or_rule"])

def build_support_qc():
    return pd.DataFrame([
        ["UNSCHED_RAW", "Visit consistency / exception routing", "Used to validate VISIT/VISITNUM context; not directly mapped to a standard LB variable"],
        ["COLL_DATE_RAW", "Datetime reconciliation", "Used to support LBDTC derivation when COLL_DTM_RAW is missing or non-ISO"],
        ["COLL_TIME_RAW", "Datetime reconciliation", "Used to support LBDTC derivation when COLL_DTM_RAW is missing or non-ISO"],
        ["VISITDT_RAW", "Operational visit-date support", "Not mapped directly in final LB in this demo; available for traceability or sponsor-specific needs"],
        ["REF_UNIT_RAW", "Range reconciliation", "Used to assess whether reference ranges can be interpreted in the same unit context as the standardized result"],
        ["RESULT_NUM_RAW", "Numeric reconciliation", "Used to support deterministic numeric interpretation of RESULT_RAW"],
        ["RESULT_CHAR_RAW", "Categorical reconciliation", "Used to support standardization of categorical results where applicable"],
        ["RESULT_QUAL_RAW", "Comparator reconciliation", "Used to interpret comparator-style results such as <0.3"],
        ["LAB_SOURCE_RAW", "Operational source tracking", "Candidate for SUPPLB or traceability use if sponsor requires"],
        ["LAB_VENDOR_RAW", "Vendor traceability", "Candidate for SUPPLB or traceability use if sponsor requires"],
        ["HEMOLYZED_RAW", "Specimen quality review", "QC/exceptions and possible future SUPPLB use"],
        ["REPEAT_RAW", "Repeat/redraw operational review", "Drives row-selection and QC/exceptions review"],
        ["SAMPLE_ID_RAW", "Sample traceability", "Candidate for SUPPLB or traceability use if sponsor requires"],
        ["COMMENT_RAW", "Operational comments", "QC/exceptions and possible future SUPPLB use"],
        ["POSTDOSE_RAW", "Timing support", "Operational timing support only in this demo; not directly mapped"],
        ["FORM_NAME", "Traceability", "Source form auditability only"],
        ["LB_PAGE_ID", "Traceability", "Source row auditability only"],
        ["LB_LINE_NO", "Traceability", "Source row auditability only"],
        ["SEX_RAW", "Reference range review", "May explain valid range variation; not directly mapped in LB"],
        ["AGE_YRS", "Reference range review", "May explain valid range variation; not directly mapped in LB"],
        ["CLIN_SIG_RAW", "Clinical review support", "Not mapped to standard LB variable in this demo; may be metadata-only or sponsor-specific"],
    ], columns=["source_variable", "purpose", "recommended_handling"])

def build_supplemental_qualifier_plan():
    return pd.DataFrame([
        ["LAB_SOURCE_RAW", "SUPPLB", "LBSRC", "Lab source", "Candidate only if sponsor wants retained operational source in submission metadata"],
        ["LAB_VENDOR_RAW", "SUPPLB", "LBVEND", "Lab vendor", "Candidate only if sponsor wants retained vendor traceability"],
        ["SAMPLE_ID_RAW", "SUPPLB", "LBSMPID", "Sample identifier", "Candidate only if sponsor wants retained sample traceability"],
        ["HEMOLYZED_RAW", "SUPPLB", "LBHEM", "Hemolyzed specimen flag", "Candidate only if sponsor decides it belongs in supplemental qualifiers"],
        ["REPEAT_RAW", "SUPPLB", "LBRPT", "Repeat/redraw flag", "Candidate only if sponsor decides it belongs in supplemental qualifiers"],
    ], columns=["raw_variable", "supp_dataset", "qnam", "qlabel", "note"])

def build_row_selection_rules():
    return pd.DataFrame([
        ["1", "Eligibility gate", "Only rows that pass Layer 1 QC and deterministic mapping requirements are eligible for final LB. Rows with unresolved contradictions remain in exception outputs."],
        ["2", "NOT DONE handling", "If NOT_DONE_RAW='Y', set LBSTAT='NOT DONE'. Keep or exclude these rows according to sponsor policy; do not populate performed-result variables on retained not-done rows."],
        ["3", "Duplicate detection key", "Potential duplicates are identified on STUDYID + USUBJID + mapped LBTESTCD + resolved LBDTC + VISITNUM."],
        ["4", "Exact duplicate handling", "If all mapped content is identical across duplicate key rows, keep one record and route the redundant copies to exception/log as dropped duplicates."],
        ["5", "Repeat/redraw preference", "When COMMENT_RAW or REPEAT_RAW indicates repeat/redraw and there is a paired earlier record for the same event/test, prefer the repeat/redraw result only if sponsor policy confirms the redraw supersedes the earlier result; otherwise route both rows for manual resolution."],
        ["6", "Multiple lab source conflict", "If the same subject/visit/test/event exists from multiple lab sources and no sponsor priority rule exists, route to exception rather than choose automatically."],
        ["7", "Hemolyzed result handling", "If HEMOLYZED_RAW='Y' with a reported result, do not auto-discard or auto-keep; follow sponsor policy or route to exception."],
        ["8", "Visit contradiction handling", "Rows with unresolved VISIT/VISITNUM/UNSCHED contradictions do not enter final LB."],
        ["9", "Range contradiction handling", "Rows with impossible reference ranges (for example low >= high) can still populate result variables if sponsor allows, but range variables should be blank and the row should be flagged in exception/QC output."],
        ["10", "Sequence assignment point", "Assign LBSEQ only after final record selection is complete."],
    ], columns=["step_no", "topic", "rule"])

def build_exception_routing():
    return pd.DataFrame([
        ["Unmapped test code/name", "Route to LB exception dataset", "No deterministic LBTESTCD/LBTEST assignment"],
        ["Specimen incompatible with mapped test", "Route to LB exception dataset", "Medically meaningful mismatch"],
        ["Invalid or contradictory datetime", "Route to LB exception dataset", "Cannot derive trustworthy LBDTC"],
        ["Unsupported numeric unit conversion", "Route to LB exception dataset", "Cannot derive trustworthy LBSTRESN/LBSTRESU"],
        ["Performed-looking row without result", "Route to LB exception dataset", "Missing required result information"],
        ["Contradictory NOT DONE logic", "Route to LB exception dataset", "Source inconsistency"],
        ["Contradictory abnormal flag vs result/range", "Route to LB exception dataset", "Cannot trust LBNRIND without review"],
        ["Duplicate or competing source records", "Route to LB exception dataset", "Requires explicit row-selection decision"],
    ], columns=["issue_pattern", "routing_action", "reason"])

def build_definexml_notes():
    return pd.DataFrame([
        ["Dataset metadata", "Dataset label, class, structure, key variables, sort order, standard/version"],
        ["Variable metadata", "Name, label, datatype, length/significant digits, origin, codelist reference, comments"],
        ["Value-level metadata", "Needed where metadata differ by LBTESTCD or result-type condition"],
        ["Codelists", "Controlled terminology should be documented for coded variables and sponsor codelists"],
        ["Origin", "Use Collected, Derived, Assigned consistently; avoid ad-hoc origin values"],
        ["Algorithms/comments", "Derived variables should carry enough algorithm detail to support define.xml and reviewer understanding"],
        ["External links", "Annotated CRF and reviewer-guide links can be attached later outside this builder script"],
    ], columns=["definexml_component", "what_to_capture"])

def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 70)

def write_excel(datasets, out_xlsx):
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    readme_rows = [
        ["Sheet", "Purpose"],
        ["Dataset_Metadata", "Dataset-level metadata for LB including class, keys, sort order, and define.xml-ready notes."],
        ["LB_Spec", "Main LB mapping specification; one row per target variable."],
        ["LB_Test_Map", "Approved raw test code/name to LBTESTCD/LBTEST/LBCAT mapping support."],
        ["LB_Unit_Map", "Original unit normalization and standard-unit planning."],
        ["LB_Visit_Map", "Visit name / visit number consistency reference."],
        ["LB_Specimen_Map", "Controlled specimen normalization support."],
        ["LB_Codelists", "Variable-level codelist planning for define.xml support."],
        ["LB_Value_Level_Metadata", "Conditions where value-level metadata are needed for define.xml."],
        ["Support_QC", "Operational raw fields used for reconciliation, traceability, and exception routing."],
        ["SUPPLB_Plan", "Candidate supplemental qualifier planning; not mapped directly unless sponsor approves."],
        ["Row_Selection_Rules", "Explicit final record-selection rules for duplicate/repeat/competing records."],
        ["Exception_Routing", "Patterns that route rows to exception outputs rather than final LB."],
        ["DefineXML_Notes", "Checklist for define.xml-ready metadata completion."],
    ]
    for r_idx, row in enumerate(readme_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_readme.cell(r_idx, c_idx).value = value
    style_sheet(ws_readme)

    for name, df in datasets:
        ws = wb.create_sheet(name)
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(1, c_idx).value = col
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx).value = value
        style_sheet(ws)

    wb.save(out_xlsx)

def main():
    OUTDIR.mkdir(exist_ok=True)

    dataset_metadata = build_dataset_metadata()
    main_spec = build_main_spec()
    test_map = build_test_map()
    unit_map = build_unit_map()
    visit_map = build_visit_map()
    specimen_map = build_specimen_map()
    codelists = build_codelists()
    vlm = build_value_level_metadata()
    support_qc = build_support_qc()
    supplb_plan = build_supplemental_qualifier_plan()
    row_selection = build_row_selection_rules()
    exception_routing = build_exception_routing()
    definexml_notes = build_definexml_notes()

    dataset_metadata.to_csv(OUTDIR / "lb_dataset_metadata_v4.csv", index=False)
    main_spec.to_csv(OUTDIR / "lb_mapping_spec_validated_v4.csv", index=False)
    test_map.to_csv(OUTDIR / "lb_test_map_v4.csv", index=False)
    unit_map.to_csv(OUTDIR / "lb_unit_map_v4.csv", index=False)
    visit_map.to_csv(OUTDIR / "lb_visit_map_v4.csv", index=False)
    specimen_map.to_csv(OUTDIR / "lb_specimen_map_v4.csv", index=False)
    codelists.to_csv(OUTDIR / "lb_codelists_v4.csv", index=False)
    vlm.to_csv(OUTDIR / "lb_value_level_metadata_v4.csv", index=False)
    support_qc.to_csv(OUTDIR / "lb_support_qc_rows_v4.csv", index=False)
    supplb_plan.to_csv(OUTDIR / "lb_supplb_plan_v4.csv", index=False)
    row_selection.to_csv(OUTDIR / "lb_row_selection_rules_v4.csv", index=False)
    exception_routing.to_csv(OUTDIR / "lb_exception_routing_v4.csv", index=False)
    definexml_notes.to_csv(OUTDIR / "lb_definexml_notes_v4.csv", index=False)

    write_excel(
        datasets=[
            ("Dataset_Metadata", dataset_metadata),
            ("LB_Spec", main_spec),
            ("LB_Test_Map", test_map),
            ("LB_Unit_Map", unit_map),
            ("LB_Visit_Map", visit_map),
            ("LB_Specimen_Map", specimen_map),
            ("LB_Codelists", codelists),
            ("LB_Value_Level_Metadata", vlm),
            ("Support_QC", support_qc),
            ("SUPPLB_Plan", supplb_plan),
            ("Row_Selection_Rules", row_selection),
            ("Exception_Routing", exception_routing),
            ("DefineXML_Notes", definexml_notes),
        ],
        out_xlsx=OUTDIR / "LB_Mapping_Spec_Industry_Grade_v4.xlsx",
    )

    readme = (
        "LB spec package v4\n"
        "- Adds explicit row-selection rules for duplicates, repeats, redraws, and competing source rows\n"
        "- Adds explicit exception-routing guidance for unmapped, contradictory, and unsupported scenarios\n"
        "- Keeps main LB variable spec focused on standard SDTM LB variables\n"
        "- Keeps define.xml support sheets for codelists and value-level metadata\n"
        "- Keeps support/QC and SUPPLB planning separate from the main LB variable spec\n"
        "- Intended as final industry-grade demo blueprint for SDTM LB generation\n"
    )
    (OUTDIR / "README.txt").write_text(readme, encoding="utf-8")
    print(f"Created outputs in: {OUTDIR}")

if __name__ == "__main__":
    main()
