import os
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

MODEL = "gpt-5.4"

DM_VARIABLE_ORDER = [
    "STUDYID", "DOMAIN", "USUBJID", "SUBJID", "SITEID",
    "RFSTDTC", "BRTHDTC", "AGE", "AGEU", "SEX",
    "RACE", "ETHNIC", "COUNTRY", "ARM", "ACTARM",
    "DTHFL", "DTHDTC",
]

DM_DOMAIN_POLICY = {
    "STUDYID": {"target_label": "Study Identifier", "target_type": "Char", "target_length": 20, "core": "Req"},
    "DOMAIN": {"target_label": "Domain Abbreviation", "target_type": "Char", "target_length": 2, "core": "Req"},
    "USUBJID": {"target_label": "Unique Subject Identifier", "target_type": "Char", "target_length": 40, "core": "Req"},
    "SUBJID": {"target_label": "Subject Identifier for the Study", "target_type": "Char", "target_length": 20, "core": "Req"},
    "SITEID": {"target_label": "Study Site Identifier", "target_type": "Char", "target_length": 10, "core": "Req"},
    "RFSTDTC": {"target_label": "Subject Reference Start Date/Time", "target_type": "Char", "target_length": 20, "core": "Exp"},
    "BRTHDTC": {"target_label": "Date/Time of Birth", "target_type": "Char", "target_length": 20, "core": "Perm"},
    "AGE": {"target_label": "Age", "target_type": "Num", "target_length": 8, "core": "Exp"},
    "AGEU": {"target_label": "Age Units", "target_type": "Char", "target_length": 10, "core": "Exp"},
    "SEX": {"target_label": "Sex", "target_type": "Char", "target_length": 1, "core": "Req"},
    "RACE": {"target_label": "Race", "target_type": "Char", "target_length": 40, "core": "Perm"},
    "ETHNIC": {"target_label": "Ethnicity", "target_type": "Char", "target_length": 40, "core": "Perm"},
    "COUNTRY": {"target_label": "Country", "target_type": "Char", "target_length": 3, "core": "Perm"},
    "ARM": {"target_label": "Description of Planned Arm", "target_type": "Char", "target_length": 40, "core": "Perm"},
    "ACTARM": {"target_label": "Description of Actual Arm", "target_type": "Char", "target_length": 40, "core": "Perm"},
    "DTHFL": {"target_label": "Subject Death Flag", "target_type": "Char", "target_length": 1, "core": "Exp"},
    "DTHDTC": {"target_label": "Date/Time of Death", "target_type": "Char", "target_length": 20, "core": "Exp"},
}

DM_VARIABLE_GUIDANCE = {
    "DOMAIN": {"allowed_origin": {"Assigned"}, "allowed_mapping_class": {"Constant"}, "reconciliation_allowed": False},
    "USUBJID": {"allowed_origin": {"Derived"}, "allowed_mapping_class": {"Derived", "Reconcile"}, "reconciliation_allowed": True, "preferred_reconciliation_source": "SUBJECT_KEY"},
    "SUBJID": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Reconcile"}, "reconciliation_allowed": False},
    "SITEID": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Reconcile"}, "reconciliation_allowed": False},
    "RFSTDTC": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "BRTHDTC": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "AGE": {"allowed_origin": {"CRF/eCRF", "Derived"}, "allowed_mapping_class": {"Direct", "Derived", "Reconcile"}, "reconciliation_allowed": True, "preferred_reconciliation_source": "DATE_OF_BIRTH + REF_START_DT"},
    "AGEU": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "SEX": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "RACE": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "ETHNIC": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "COUNTRY": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "ARM": {"allowed_origin": {"CRF/eCRF", "Assigned"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "ACTARM": {"allowed_origin": {"CRF/eCRF", "Assigned"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "DTHFL": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "DTHDTC": {"allowed_origin": {"CRF/eCRF"}, "allowed_mapping_class": {"Direct", "Recode"}, "reconciliation_allowed": False},
    "STUDYID": {"allowed_origin": {"CRF/eCRF", "Assigned"}, "allowed_mapping_class": {"Direct", "Constant"}, "reconciliation_allowed": False},
}

PROMPT_TEMPLATE = '''
You are a senior clinical standards programmer drafting a production-style SDTM DM mapping specification.

Your job in this step:
1. reasoning quality
2. terminology distinctions
3. spec-writing style
4. correct semantics

A downstream validator will enforce domain conformance, impossible combinations, and consistency checks.

Rules:
- Only use these DM target variables: __ALLOWED_DM_VARIABLES__.
- Use reviewable placeholders like "Not provided in source metadata" instead of blanks.
- Preserve identifiers as character data and preserve left-padding where applicable.
- Preserve partial dates; do not silently impute.
- Do not mark something as Derived merely because you changed representation.
- Converting collected dates to ISO 8601 is usually Recode, not Derived.
- Populate reconciliation_source only when another variable is truly used to verify or reconcile the target.
- Make the spec explicit enough that deterministic transformation code could be written from it.

Important distinctions:
Origin:
- CRF/eCRF = directly collected
- Assigned = constant/externally assigned value
- Derived = computed from one or more variables
- Support/QC only = not used to populate final target

Mapping Class:
- Direct = copied without conceptual derivation
- Recode = same concept, standardized representation/coding
- Derived = computed from one or more variables
- Constant = fixed assigned value
- Reconcile = primary mapping plus cross-check
- Support/QC only = support-only field

DM-specific guidance:
- DOMAIN is Assigned + Constant and source_variable should be "N/A".
- USUBJID is commonly Derived from component identifiers.
- SUBJID and SITEID are generally direct collected identifiers, not Derived.
- If collected AGE exists, AGE should generally be Direct, not Derived.
- DTHFL should follow SDTM convention "Y or null", not "Y/N".
- BRTHDTC, RFSTDTC, and DTHDTC are typically CRF/eCRF with Direct or Recode when converted to ISO 8601 character representation.
- COUNTRY should be explicit about target standard if target length is 3.
- SEX rule should explicitly handle unknown as U.

Return JSON ONLY with:
{
  "study_assumptions": ["string"],
  "mapping_rows": [
    {
      "source_form_or_module": "string or null",
      "source_variable": "string",
      "source_label": "string or null",
      "source_type_hint": "string or null",
      "sample_values": ["string"],
      "target_domain": "DM or null",
      "target_variable": "allowed DM variable or null",
      "target_label": "string or null",
      "target_type": "Char | Num | null",
      "target_length": "integer or null",
      "core": "Req | Exp | Perm | null",
      "origin": "CRF/eCRF | Assigned | Derived | Support/QC only",
      "mapping_class": "Direct | Constant | Recode | Derived | Reconcile | Support/QC only",
      "source_role_in_rule": "Primary | Secondary | Reconciliation | Support | null",
      "reconciliation_source": "string or null",
      "rule": "string",
      "controlled_terms_or_format": "string or null",
      "primary_qc_checks": "string",
      "traceability_note": "string or null",
      "programming_notes": "string or null",
      "review_status": "Draft",
      "confidence": "High | Medium | Low",
      "ambiguity_note": "string or null"
    }
  ]
}
'''

def profile_source(df: pd.DataFrame, max_samples: int = 5) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    force_char_if_name_contains = ["ID", "NUMBER", "KEY", "CODE"]
    date_like_names = ["DT", "DTC", "DATE"]
    for col in df.columns:
        series = df[col]
        non_null = series.dropna().astype(str).str.strip()
        non_null = non_null[non_null != ""]
        unique_samples = list(dict.fromkeys(non_null.head(20).tolist()))[:max_samples]
        upper_col = col.upper()
        if any(token in upper_col for token in force_char_if_name_contains):
            type_hint = "char"
        elif any(token in upper_col for token in date_like_names):
            type_hint = "date-like char"
        else:
            type_hint = "char"
            if len(non_null) > 0:
                coerced = pd.to_numeric(non_null, errors="coerce")
                if coerced.notna().mean() > 0.8:
                    type_hint = "numeric"
        rows.append({
            "source_form_or_module": "Not provided in source metadata",
            "source_variable": col,
            "source_label": "Not provided in source metadata",
            "source_type_hint": type_hint,
            "sample_values": unique_samples,
            "non_null_count": int(non_null.shape[0]),
            "distinct_count": int(non_null.nunique()),
        })
    return rows

def build_input_payload(source_profile: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {"domain": "DM", "allowed_target_variables": DM_VARIABLE_ORDER, "source_profile": source_profile}

def fallback_rows() -> Dict[str, Any]:
    source_map = {
        "STUDYID": "PROTOCOL_ID", "DOMAIN": "N/A", "USUBJID": "PROTOCOL_ID + SITE_NUMBER + SUBJECT_NUMBER",
        "SUBJID": "SUBJECT_NUMBER", "SITEID": "SITE_NUMBER", "RFSTDTC": "REF_START_DT",
        "BRTHDTC": "DATE_OF_BIRTH", "AGE": "AGE_AT_REF", "AGEU": "AGE_UNITS", "SEX": "SEX_AT_BIRTH",
        "RACE": "RACE_CAT", "ETHNIC": "ETHNIC_GRP", "COUNTRY": "COUNTRY_CODE",
        "ARM": "PLANNED_TRT_ARM", "ACTARM": "ACTUAL_TRT_ARM", "DTHFL": "DEATH_IND", "DTHDTC": "DEATH_DATE"
    }
    rule_map = {
        "DOMAIN": 'Set DOMAIN = "DM" for all records.',
        "USUBJID": 'Derive USUBJID as PROTOCOL_ID || "-" || SITE_NUMBER || "-" || SUBJECT_NUMBER; preserve SITE_NUMBER left-padding. If SUBJECT_KEY exists, use it only for reconciliation.',
        "COUNTRY": 'Standardize COUNTRY_CODE to ISO 3166-1 alpha-3.',
        "SEX": 'If SEX_AT_BIRTH in ("Male","M","male","m") then SEX="M"; if in ("Female","F","female","f") then SEX="F"; else if in ("Unknown","U","unknown","u") then SEX="U"; blank remains blank for review.',
        "DTHFL": 'If death indicator implies death then set DTHFL="Y"; otherwise leave null. Do not populate "N" in SDTM.'
    }
    rows = []
    for var in DM_VARIABLE_ORDER:
        rows.append({
            "source_form_or_module": "Not collected / Assigned" if var == "DOMAIN" else "Not provided in source metadata",
            "source_variable": source_map[var],
            "source_label": "N/A" if var == "DOMAIN" else "Not provided in source metadata",
            "source_type_hint": None,
            "sample_values": [],
            "target_domain": "DM",
            "target_variable": var,
            "target_label": DM_DOMAIN_POLICY[var]["target_label"],
            "target_type": DM_DOMAIN_POLICY[var]["target_type"],
            "target_length": DM_DOMAIN_POLICY[var]["target_length"],
            "core": DM_DOMAIN_POLICY[var]["core"],
            "origin": "Assigned" if var == "DOMAIN" else ("Derived" if var == "USUBJID" else "CRF/eCRF"),
            "mapping_class": "Constant" if var == "DOMAIN" else ("Derived" if var == "USUBJID" else ("Direct" if var in {"STUDYID","SUBJID","SITEID","AGE","ARM","ACTARM"} else "Recode")),
            "source_role_in_rule": "Support" if var == "DOMAIN" else "Primary",
            "reconciliation_source": "SUBJECT_KEY" if var == "USUBJID" else ("DATE_OF_BIRTH + REF_START_DT" if var == "AGE" else ""),
            "rule": rule_map.get(var, f"Draft mapping for {var}; review required."),
            "controlled_terms_or_format": "ISO 3166-1 alpha-3" if var == "COUNTRY" else ("M/F/U" if var == "SEX" else "None."),
            "primary_qc_checks": "Review required.",
            "traceability_note": "Generated from fallback mode; review required.",
            "programming_notes": "Fallback mode.",
            "review_status": "Draft",
            "confidence": "Medium",
            "ambiguity_note": "AI call unavailable or failed; fallback draft created."
        })
    return {"study_assumptions": ["Fallback mode used because AI draft was unavailable."], "mapping_rows": rows}

def call_openai_for_spec(payload: Dict[str, Any]) -> Tuple[Dict[str, Any], str]:
    if OpenAI is None:
        return fallback_rows(), "openai package not available; used fallback"
    if not os.environ.get("OPENAI_API_KEY"):
        return fallback_rows(), "OPENAI_API_KEY not set; used fallback"
    client = OpenAI()
    prompt = PROMPT_TEMPLATE.replace("__ALLOWED_DM_VARIABLES__", ", ".join(DM_VARIABLE_ORDER))
    try:
        response = client.responses.create(
            model=MODEL,
            input=[
                {"role": "developer", "content": [{"type": "input_text", "text": prompt}]},
                {"role": "user", "content": [{"type": "input_text", "text": json.dumps(payload, indent=2)}]},
            ],
        )
        text = getattr(response, "output_text", None)
        if not text:
            return fallback_rows(), "AI response had no output_text; used fallback"
        return json.loads(text), "AI draft succeeded"
    except Exception as e:
        return fallback_rows(), f"AI call failed: {type(e).__name__}: {e}"

def _fill_placeholder(v: Any, default: str) -> str:
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default

def validate_and_normalize(result: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.DataFrame(result["mapping_rows"]).copy()
    issues = []

    for col, default in {
        "source_form_or_module": "Not provided in source metadata",
        "source_label": "Not provided in source metadata",
        "traceability_note": "Traceability to be confirmed during review.",
        "programming_notes": "None.",
        "controlled_terms_or_format": "None.",
        "ambiguity_note": "None.",
        "reconciliation_source": "",
    }.items():
        df[col] = df[col].apply(lambda x: _fill_placeholder(x, default))

    for var, meta in DM_DOMAIN_POLICY.items():
        mask = df["target_variable"] == var
        for key, val in meta.items():
            observed_series = df.loc[mask, key]
            for idx in observed_series.index:
                if str(df.at[idx, key]) != str(val):
                    issues.append({
                        "row_index": int(idx), "target_variable": var, "issue_type": "DOMAIN_CONFORMANCE",
                        "field": key, "severity": "WARNING", "observed": df.at[idx, key], "expected": val,
                        "action": "normalized to domain policy"
                    })
            df.loc[mask, key] = val

    preferred_mapping = {
        "DOMAIN": "Constant", "USUBJID": "Derived", "SUBJID": "Direct", "SITEID": "Direct",
        "RFSTDTC": "Recode", "BRTHDTC": "Recode", "AGE": "Direct", "AGEU": "Recode",
        "SEX": "Recode", "RACE": "Recode", "ETHNIC": "Recode", "COUNTRY": "Recode",
        "ARM": "Direct", "ACTARM": "Direct", "DTHFL": "Recode", "DTHDTC": "Recode", "STUDYID": "Direct",
    }
    preferred_origin = {"DOMAIN": "Assigned", "USUBJID": "Derived"}

    for idx, row in df.iterrows():
        var = row.get("target_variable")
        if pd.isna(var) or var is None:
            continue
        var = str(var)
        guide = DM_VARIABLE_GUIDANCE.get(var, {})

        if "allowed_origin" in guide and row["origin"] not in guide["allowed_origin"]:
            issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "IMPOSSIBLE_COMBINATION", "field": "origin", "severity": "WARNING", "observed": row["origin"], "expected": ", ".join(sorted(guide["allowed_origin"])), "action": "reset to preferred"})
            df.at[idx, "origin"] = preferred_origin.get(var, "CRF/eCRF")

        if "allowed_mapping_class" in guide and row["mapping_class"] not in guide["allowed_mapping_class"]:
            issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "IMPOSSIBLE_COMBINATION", "field": "mapping_class", "severity": "WARNING", "observed": row["mapping_class"], "expected": ", ".join(sorted(guide["allowed_mapping_class"])), "action": "reset to preferred"})
            df.at[idx, "mapping_class"] = preferred_mapping.get(var, row["mapping_class"])

        if var == "DOMAIN":
            df.at[idx, "source_variable"] = "N/A"
            df.at[idx, "source_form_or_module"] = "Not collected / Assigned"
            df.at[idx, "source_label"] = "N/A"
            df.at[idx, "source_type_hint"] = "N/A"
            df.at[idx, "source_role_in_rule"] = "Support"

        reconciliation_allowed = guide.get("reconciliation_allowed", False)
        if not reconciliation_allowed and str(row.get("reconciliation_source", "")).strip():
            issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "CONSISTENCY_CHECK", "field": "reconciliation_source", "severity": "INFO", "observed": row["reconciliation_source"], "expected": "", "action": "cleared"})
            df.at[idx, "reconciliation_source"] = ""
        elif reconciliation_allowed and not str(df.at[idx, "reconciliation_source"]).strip():
            preferred = guide.get("preferred_reconciliation_source", "")
            if preferred:
                issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "CONSISTENCY_CHECK", "field": "reconciliation_source", "severity": "INFO", "observed": "", "expected": preferred, "action": "filled preferred reconciliation source"})
                df.at[idx, "reconciliation_source"] = preferred

        if str(row.get("review_status", "")) != "Draft":
            issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "RULE_ENFORCEMENT", "field": "review_status", "severity": "INFO", "observed": row.get("review_status"), "expected": "Draft", "action": "reset"})
            df.at[idx, "review_status"] = "Draft"

        if var == "COUNTRY":
            target_len = str(df.at[idx, "target_length"])
            ctf = str(df.at[idx, "controlled_terms_or_format"])
            if target_len == "3" and "ISO" not in ctf.upper():
                issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "RULE_ENFORCEMENT", "field": "controlled_terms_or_format", "severity": "WARNING", "observed": ctf, "expected": "ISO 3166-1 alpha-3", "action": "set explicit target standard"})
                df.at[idx, "controlled_terms_or_format"] = "ISO 3166-1 alpha-3"

        if var == "SEX":
            rule_text = str(df.at[idx, "rule"])
            if '"U"' not in rule_text and " U " not in rule_text:
                issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "RULE_ENFORCEMENT", "field": "rule", "severity": "WARNING", "observed": rule_text, "expected": "explicit U mapping", "action": "leave for human review"})
        if var == "USUBJID":
            src_var = str(df.at[idx, "source_variable"])
            rule_text = str(df.at[idx, "rule"])
            if src_var == "SUBJECT_KEY" and "PROTOCOL_ID" in rule_text and "SITE_NUMBER" in rule_text and "SUBJECT_NUMBER" in rule_text:
                issues.append({"row_index": int(idx), "target_variable": var, "issue_type": "CONSISTENCY_CHECK", "field": "source_variable", "severity": "WARNING", "observed": src_var, "expected": "component identifiers or explicit reconcile pattern", "action": "leave for human review"})

    allowed_reconcile_vars = {"AGE", "USUBJID"}
    non_allowed_mask = df["target_variable"].notna() & ~df["target_variable"].isin(allowed_reconcile_vars)
    for idx in df.index[non_allowed_mask]:
        df.at[idx, "reconciliation_source"] = ""

    df.loc[df["target_variable"].notna(), "target_domain"] = "DM"
    df["sample_values_joined"] = df["sample_values"].apply(lambda x: " | ".join(x) if isinstance(x, list) else "")
    order_map = {var: i for i, var in enumerate(DM_VARIABLE_ORDER, start=1)}
    df["target_order"] = df["target_variable"].map(order_map).fillna(999).astype(int)

    mapped_df = df[df["target_variable"].notna()].copy().sort_values(["target_order", "source_variable"]).reset_index(drop=True)
    mapped_df.insert(0, "spec_seq", range(1, len(mapped_df) + 1))

    support_df = df[df["target_variable"].isna()].copy().sort_values(["source_variable"]).reset_index(drop=True)
    support_df.insert(0, "support_seq", range(1, len(support_df) + 1))

    issues_df = pd.DataFrame(issues)
    if not issues_df.empty:
        issues_df = issues_df.sort_values(["severity", "target_variable", "field"]).reset_index(drop=True)
    return mapped_df, support_df, issues_df

def compare_with_prior(current_df: pd.DataFrame, prior_csv: Path) -> pd.DataFrame:
    if not prior_csv.exists():
        return pd.DataFrame(columns=["target_variable", "change_type", "field", "old_value", "new_value"])
    old = pd.read_csv(prior_csv, dtype=str)
    rename_map = {
        "Target Variable": "target_variable",
        "Source Variable": "source_variable",
        "Origin": "origin",
        "Mapping Class": "mapping_class",
        "Reconciliation Source": "reconciliation_source",
        "Transformation / Derivation Rule": "rule",
        "Core": "core",
        "Type": "target_type",
        "Length": "target_length",
    }
    old = old.rename(columns={k: v for k, v in rename_map.items() if k in old.columns})
    keep = ["target_variable", "source_variable", "origin", "mapping_class", "reconciliation_source", "rule", "core", "target_type", "target_length"]
    old = old[[c for c in keep if c in old.columns]].copy()
    new = current_df[[c for c in keep if c in current_df.columns]].copy()

    changes = []
    old_map = {str(r["target_variable"]): r for _, r in old.iterrows()} if "target_variable" in old.columns else {}
    new_map = {str(r["target_variable"]): r for _, r in new.iterrows()}

    for var, new_row in new_map.items():
        if var not in old_map:
            changes.append({"target_variable": var, "change_type": "ADDED", "field": "", "old_value": "", "new_value": ""})
            continue
        old_row = old_map[var]
        for field in keep[1:]:
            ov = str(old_row.get(field, "") or "")
            nv = str(new_row.get(field, "") or "")
            if ov != nv:
                changes.append({"target_variable": var, "change_type": "MODIFIED", "field": field, "old_value": ov, "new_value": nv})

    for var in old_map:
        if var not in new_map:
            changes.append({"target_variable": var, "change_type": "REMOVED", "field": "", "old_value": "", "new_value": ""})

    return pd.DataFrame(changes)

def write_outputs(result: Dict[str, Any], mapped_df: pd.DataFrame, support_df: pd.DataFrame, issues_df: pd.DataFrame, changes_df: pd.DataFrame, output_dir: Path, status_text: str):
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "dm_mapping_spec_ai_raw_v5.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
    mapped_df.to_csv(output_dir / "dm_mapping_spec_validated_v5.csv", index=False)
    support_df.to_csv(output_dir / "dm_support_qc_rows_v5.csv", index=False)
    issues_df.to_csv(output_dir / "dm_validator_issues_v5.csv", index=False)
    changes_df.to_csv(output_dir / "dm_change_log_v5.csv", index=False)
    (output_dir / "run_status.txt").write_text(status_text, encoding="utf-8")
    (output_dir / "dm_study_assumptions_v5.txt").write_text("\n".join([f"- {x}" for x in result.get("study_assumptions", [])]), encoding="utf-8")
    (output_dir / "dm_lifecycle_notes.txt").write_text(
        "Lifecycle guidance:\n"
        "- If the approved spec changes, teams typically update code and regenerate the affected SDTM domain from source.\n"
        "- Controlled rebuilds replace the previous domain build for the next snapshot/release.\n"
        "- Change logs help identify impact, but controlled rerun is the standard path.\n",
        encoding="utf-8"
    )

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_spec = wb.create_sheet("DM_Spec")
    ws_support = wb.create_sheet("Support_QC")
    ws_val = wb.create_sheet("Validator_Issues")
    ws_changes = wb.create_sheet("Change_Log")

    for r, row in enumerate([
        ["Layer", "Responsibility"],
        ["Prompt / AI", "Reasoning quality, terminology distinctions, spec-writing style, correct semantics."],
        ["Validator", "Rule enforcement, domain conformance, impossible combinations, consistency checks."],
        ["Human", "Sponsor-specific nuance, study-specific decisions, final approval."],
        ["Run status", status_text],
    ], start=1):
        for c, val in enumerate(row, start=1):
            ws_readme.cell(r, c).value = val

    spec_headers = ["Seq","Domain","Target Variable","Target Label","Core","Type","Length","Source Form / Module","Source Variable","Source Label","Source Type Hint","Sample Values","Origin","Mapping Class","Source Role in Rule","Reconciliation Source","Transformation / Derivation Rule","Controlled Terms / Format","Primary QC Checks","Traceability Note","Programming Notes","Review Status","AI Confidence","Ambiguity Note"]
    spec_cols = ["spec_seq","target_domain","target_variable","target_label","core","target_type","target_length","source_form_or_module","source_variable","source_label","source_type_hint","sample_values_joined","origin","mapping_class","source_role_in_rule","reconciliation_source","rule","controlled_terms_or_format","primary_qc_checks","traceability_note","programming_notes","review_status","confidence","ambiguity_note"]

    for c, h in enumerate(spec_headers, start=1):
        ws_spec.cell(1, c).value = h
    for r_idx, row in enumerate(mapped_df[spec_cols].itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            if spec_cols[c_idx-1] == "target_length" and value not in (None, ""):
                try:
                    value = int(float(value))
                except Exception:
                    pass
            ws_spec.cell(r_idx, c_idx).value = value

    support_headers = ["Seq","Source Form / Module","Source Variable","Source Label","Source Type Hint","Sample Values","Origin","Mapping Class","Source Role in Rule","Reconciliation Source","Rule","Primary QC Checks","Traceability Note","Programming Notes","AI Confidence","Ambiguity Note"]
    support_cols = ["support_seq","source_form_or_module","source_variable","source_label","source_type_hint","sample_values_joined","origin","mapping_class","source_role_in_rule","reconciliation_source","rule","primary_qc_checks","traceability_note","programming_notes","confidence","ambiguity_note"]
    for c, h in enumerate(support_headers, start=1):
        ws_support.cell(1, c).value = h
    for r_idx, row in enumerate(support_df[support_cols].itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_support.cell(r_idx, c_idx).value = value

    if issues_df.empty:
        ws_val["A1"] = "No validator issues."
    else:
        for c_idx, h in enumerate(list(issues_df.columns), start=1):
            ws_val.cell(1, c_idx).value = h
        for r_idx, row in enumerate(issues_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_val.cell(r_idx, c_idx).value = value

    if changes_df.empty:
        ws_changes["A1"] = "No prior-spec differences found."
    else:
        for c_idx, h in enumerate(list(changes_df.columns), start=1):
            ws_changes.cell(1, c_idx).value = h
        for r_idx, row in enumerate(changes_df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_changes.cell(r_idx, c_idx).value = value

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    for ws in [ws_readme, ws_spec, ws_support, ws_val, ws_changes]:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_readme.column_dimensions["A"].width = 18
    ws_readme.column_dimensions["B"].width = 100
    widths = {"A":6,"B":8,"C":16,"D":28,"E":8,"F":10,"G":8,"H":22,"I":20,"J":22,"K":16,"L":26,"M":14,"N":14,"O":16,"P":22,"Q":58,"R":24,"S":34,"T":28,"U":26,"V":12,"W":12,"X":30}
    for col, width in widths.items():
        ws_spec.column_dimensions[col].width = width
    for col, width in {"A":6,"B":22,"C":20,"D":20,"E":14,"F":24,"G":14,"H":16,"I":16,"J":22,"K":50,"L":28,"M":26,"N":24,"O":12,"P":30}.items():
        ws_support.column_dimensions[col].width = width
    for ws in [ws_val, ws_changes]:
        for col in "ABCDEFGH":
            ws.column_dimensions[col].width = 20

    ws_spec.freeze_panes = "A2"
    ws_support.freeze_panes = "A2"
    ws_val.freeze_panes = "A2"
    ws_changes.freeze_panes = "A2"

    wb.save(output_dir / "DM_Mapping_Spec_Industry_Grade_v5.xlsx")

def main():
    script_dir = Path(__file__).resolve().parent
    source_csv = script_dir / "dm_raw_crf_style_50_rows.csv"
    prior_spec_csv = script_dir / "dm_mapping_spec_validated_v4.csv"
    output_dir = script_dir / "ai_dm_spec_outputs_v5"

    if not source_csv.exists():
        raise FileNotFoundError(f"Source file not found: {source_csv}")

    df = pd.read_csv(source_csv, dtype=str)
    payload = build_input_payload(profile_source(df))
    result, status_text = call_openai_for_spec(payload)
    mapped_df, support_df, issues_df = validate_and_normalize(result)
    changes_df = compare_with_prior(mapped_df, prior_spec_csv)
    write_outputs(result, mapped_df, support_df, issues_df, changes_df, output_dir, status_text)
    print(f"Created outputs in: {output_dir}")

if __name__ == "__main__":
    main()
