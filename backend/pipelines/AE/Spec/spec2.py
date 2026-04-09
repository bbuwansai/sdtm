from pathlib import Path
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).resolve().parent
OUTDIR = BASE / "ae_spec_outputs_v2"

EXPECTED_RAW_COLUMNS = [
    "ROW_ID","STUDYID_RAW","SITEID_RAW","SUBJECT_RAW","SCREENING_NO","RAND_NO","VISIT_RAW","VISITDT_RAW",
    "AE_FORM_SEQ","AE_SEQ_CRf","AEYN_RAW","AE_TERM","AE_START_DATE_RAW","AE_START_TIME_RAW","AE_END_DATE_RAW",
    "AE_END_TIME_RAW","AE_ONGOING_RAW","AE_SEVERITY_RAW","AE_TOXGR_RAW","AE_SER_RAW","AE_SER_DTH_RAW",
    "AE_SER_LIFE_RAW","AE_SER_HOSP_RAW","AE_SER_DISAB_RAW","AE_SER_CONG_RAW","AE_SER_MIE_RAW",
    "AE_REL_STUDY_DRUG_RAW","AE_REL_STUDY_DRUG2_RAW","AE_ACTION_DRUG_RAW","AE_ACTION_DRUG2_RAW",
    "AE_ACTION_OTHER_TXT","AE_OUTCOME_RAW","AE_PRESPEC_RAW","AE_REPORTED_BY","AE_REPORT_DATE_RAW",
    "AE_COMMENT","ENTRY_STATUS_RAW","CHANGE_REASON_RAW"
]


def auto_detect_file(candidates):
    for cand in candidates:
        p = BASE / cand
        if p.exists():
            return p
    return None


def read_json(path):
    if path is None or not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def build_dataset_metadata():
    return pd.DataFrame([
        {
            "dataset_name": "AE",
            "label": "Adverse Events",
            "class": "Events",
            "structure": "One record per adverse event per subject in the submitted AE dataset. Operational CRF rows may be collapsed to one final event record or retained as separate records based on sponsor safety-review rules.",
            "domain_keys": "STUDYID, USUBJID, AESEQ",
            "sort_order": "STUDYID, USUBJID, AESTDTC, AEDECOD, AESEQ",
            "standard": "SDTMIG v3.4-aligned implementation design",
            "controlled_terminology_version": "To be populated with sponsor-approved CDISC CT package version",
            "dictionary_name": "MedDRA",
            "dictionary_version": "To be populated with sponsor-approved MedDRA version used for coding",
            "implementation_note": "This spec enforces CDISC SDTMIG compliance, MedDRA coding dependency, controlled terminology adherence, define.xml-ready metadata structure, exception routing, SUPPAE planning, and reviewer traceability.",
            "submission_gate_note": "Final coded AE submission dataset must not contain uncoded events. Rows without completed coding or with unresolved contradictions remain outside final submission output until resolved.",
            "final_record_selection_rule": "Only rows passing Layer 1 QC, coding dependency checks, and deterministic mapping rules enter candidate AE build. Rows with unresolved contradictions remain in exception outputs. Submission structure can either summarize operational rows into one final event per subject/event or preserve separate records when seriousness, severity, or causality changes are intended to define distinct events."
        }
    ])


def build_main_spec():
    rows = [
        {
            "spec_seq": 1, "target_domain": "AE", "target_variable": "STUDYID",
            "target_label": "Study Identifier", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 20,
            "core": "Req", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "STUDYID_RAW",
            "source_label": "Raw Study Identifier", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "", "codelist_ref": "",
            "rule": "Map STUDYID = strip(STUDYID_RAW). Preserve the collected study identifier as character.",
            "primary_qc_checks": "STUDYID_RAW must not be missing and should be constant across the study.",
            "traceability_note": "Direct trace to source field STUDYID_RAW.",
            "programming_notes": "Do not cast to numeric.",
            "definexml_origin_comment": "Collected on CRF.",
            "exception_condition": "Missing or inconsistent STUDYID_RAW -> exception",
            "final_condition": "Non-missing and study-consistent", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 2, "target_domain": "AE", "target_variable": "DOMAIN",
            "target_label": "Domain Abbreviation", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 2,
            "core": "Req", "origin": "Assigned", "mapping_class": "Constant",
            "source_form_or_module": "Not collected / Assigned", "source_variable": "Not Applicable",
            "source_label": "Not Applicable", "source_type_hint": "N/A", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "", "codelist_ref": "",
            "rule": 'Set DOMAIN = "AE" for all records.',
            "primary_qc_checks": 'DOMAIN must equal "AE" for every output row.',
            "traceability_note": "Assigned constant for AE domain build.",
            "programming_notes": "No source field is used.",
            "definexml_origin_comment": "Assigned constant.",
            "exception_condition": 'If DOMAIN is not "AE" -> exception',
            "final_condition": 'Always "AE"', "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 3, "target_domain": "AE", "target_variable": "USUBJID",
            "target_label": "Unique Subject Identifier", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 40,
            "core": "Req", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "STUDYID_RAW + SITEID_RAW + SUBJECT_RAW",
            "source_label": "Study + Site + Subject", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "AE_USUBJID_01",
            "reconciliation_source": "SCREENING_NO + RAND_NO", "codelist_ref": "",
            "rule": "Derive USUBJID = STUDYID_RAW || '-' || SITEID_RAW || '-' || SUBJECT_RAW, preserving site and subject identifiers as character strings. If sponsor already uses this exact collected format, preserve it.",
            "primary_qc_checks": "USUBJID must not be missing. Final AE output must not contain duplicate STUDYID/USUBJID/AESEQ combinations.",
            "traceability_note": "Traceable to study, site, and subject identifying fields in raw AE.",
            "programming_notes": "Do not coerce SITEID_RAW or SUBJECT_RAW to numeric if left-padding may be lost.",
            "definexml_origin_comment": "Derived from collected identifiers.",
            "exception_condition": "Missing or duplicate subject components -> exception",
            "final_condition": "Unique and non-missing", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 4, "target_domain": "AE", "target_variable": "AESEQ",
            "target_label": "Sequence Number", "sdtm_role": "Identifier", "target_type": "Num", "target_length": 8,
            "core": "Req", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Assigned", "source_variable": "Sorted final AE record order",
            "source_label": "Operational sequence", "source_type_hint": "numeric", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "AE_AESEQ_01",
            "reconciliation_source": "AE_SEQ_CRf + AE_FORM_SEQ + ROW_ID", "codelist_ref": "",
            "rule": "After final AE record selection and sorting by STUDYID, USUBJID, AESTDTC, AEDECOD (or AETERM when uncoded interim review is permitted outside final submission), assign AESEQ as a sequential integer starting at 1 within each USUBJID. AE_SEQ_CRf is a traceability input only and is not trusted as final AESEQ when duplicates or gaps exist.",
            "primary_qc_checks": "AESEQ must be non-missing, numeric, and unique within USUBJID.",
            "traceability_note": "Derived operational sequence for final SDTM AE; raw AE_SEQ_CRf retained for traceability and reconciliation.",
            "programming_notes": "Assign after record selection, coding decision, and exception handling. Never trust raw AE_SEQ_CRf as final sequence without review.",
            "definexml_origin_comment": "Derived by algorithm after final row selection.",
            "exception_condition": "Duplicate or missing AESEQ -> exception",
            "final_condition": "Unique within subject", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 5, "target_domain": "AE", "target_variable": "AETERM",
            "target_label": "Reported Term for the Adverse Event", "sdtm_role": "Topic", "target_type": "Char", "target_length": 200,
            "core": "Req", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_TERM",
            "source_label": "Verbatim Adverse Event Term", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "CHANGE_REASON_RAW + ENTRY_STATUS_RAW", "codelist_ref": "",
            "rule": "Map AETERM = trimmed AE_TERM preserving investigator-reported wording. Correct obvious whitespace only; do not auto-clinically normalize the text here.",
            "primary_qc_checks": "AETERM must not be missing when AEYN_RAW indicates an event record should exist.",
            "traceability_note": "Direct trace to verbatim AE term captured on CRF.",
            "programming_notes": "Do not apply medical coding logic directly in AETERM.",
            "definexml_origin_comment": "Collected verbatim event term.",
            "exception_condition": "Missing verbatim term on AE-present row -> exception",
            "final_condition": "Non-missing trimmed text", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 6, "target_domain": "AE", "target_variable": "AEMODIFY",
            "target_label": "Modified Reported Term", "sdtm_role": "Qualifier", "target_type": "Char", "target_length": 200,
            "core": "Perm", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Adverse Events CRF / Coding workflow", "source_variable": "AE_TERM + coding review decisions",
            "source_label": "Modified term when needed for coding", "source_type_hint": "char", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "", "value_level_condition": "Populate only when the reported term is modified to facilitate coding", "method_id": "AE_AEMODIFY_01",
            "reconciliation_source": "CHANGE_REASON_RAW + coding review log", "codelist_ref": "",
            "rule": "Leave AEMODIFY blank by default. Populate only when sponsor-approved coding review intentionally changes the reported text to facilitate MedDRA coding, while retaining AETERM as originally reported.",
            "primary_qc_checks": "If AEMODIFY is populated, coding documentation should justify the modification.",
            "traceability_note": "Traceable to coding review decision, not directly to a single raw field.",
            "programming_notes": "Do not populate for simple trimming or case normalization alone.",
            "definexml_origin_comment": "Derived during coding workflow when sponsor-approved term modification is required.",
            "exception_condition": "Unsupported or undocumented term modification -> exception/review",
            "final_condition": "Blank or documented modified term", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Depends on sponsor coding governance."
        },
        {
            "spec_seq": 7, "target_domain": "AE", "target_variable": "AEDECOD",
            "target_label": "Dictionary-Derived Term", "sdtm_role": "Result Qualifier", "target_type": "Char", "target_length": 200,
            "core": "Req", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Coding workflow", "source_variable": "AETERM or AEMODIFY",
            "source_label": "Coded preferred term", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "MedDRA Preferred Term", "value_level_condition": "", "method_id": "AE_MEDDRA_01",
            "reconciliation_source": "Coding dictionary version + coder review log", "codelist_ref": "MedDRA_PT",
            "rule": "Map AEDECOD strictly from validated MedDRA coding output. Final SDTM submission dataset must not contain uncoded AE records. Transformation logic must never attempt to infer or guess coding.",
            "primary_qc_checks": "AEDECOD must be non-missing in the submission dataset. Missing coding must be routed to coding workflow or exception handling before final output.",
            "traceability_note": "Traceable to external MedDRA coding output.",
            "programming_notes": "Separate coding dependency from raw-to-SDTM transformation logic.",
            "definexml_origin_comment": "Assigned from MedDRA coding; dictionary name and version should be captured in define.xml external codelist metadata.",
            "exception_condition": "Missing or ambiguous coding -> exception/coding queue",
            "final_condition": "Mapped to approved MedDRA PT", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 8, "target_domain": "AE", "target_variable": "AESOC",
            "target_label": "Primary System Organ Class", "sdtm_role": "Variable Qualifier", "target_type": "Char", "target_length": 100,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Coding workflow", "source_variable": "MedDRA coding output",
            "source_label": "Primary SOC", "source_type_hint": "char", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "MedDRA SOC", "value_level_condition": "", "method_id": "AE_MEDDRA_01",
            "reconciliation_source": "AEDECOD", "codelist_ref": "MedDRA_SOC",
            "rule": "Populate AESOC from sponsor-approved MedDRA coding output using the primary path associated with the coded term.",
            "primary_qc_checks": "If coding is complete, AESOC should be consistent with AEDECOD and the selected MedDRA version.",
            "traceability_note": "Traceable to MedDRA coding output.",
            "programming_notes": "Do not hardcode SOC names in the transformation script.",
            "definexml_origin_comment": "Assigned from MedDRA coding output.",
            "exception_condition": "Coding incomplete or inconsistent -> exception/coding review",
            "final_condition": "Blank or coded SOC in final submission extract", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 9, "target_domain": "AE", "target_variable": "AEBODSYS",
            "target_label": "Body System or Organ Class", "sdtm_role": "Variable Qualifier", "target_type": "Char", "target_length": 100,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Coding workflow", "source_variable": "MedDRA coding output",
            "source_label": "Body system / SOC", "source_type_hint": "char", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "MedDRA SOC", "value_level_condition": "", "method_id": "AE_MEDDRA_01",
            "reconciliation_source": "AEDECOD + AESOC", "codelist_ref": "MedDRA_SOC",
            "rule": "Populate AEBODSYS from sponsor-approved coding output. In standard sponsor implementations this is typically the same primary SOC text used for AESOC.",
            "primary_qc_checks": "If populated, AEBODSYS should align with the MedDRA coding output for the selected term.",
            "traceability_note": "Traceable to MedDRA coding output.",
            "programming_notes": "Keep AESOC/AEBODSYS handling explicitly documented; some sponsors carry both for reviewer convenience.",
            "definexml_origin_comment": "Assigned from MedDRA coding output.",
            "exception_condition": "Coding incomplete or inconsistent -> exception/coding review",
            "final_condition": "Blank or coded body system", "review_status": "Final", "confidence": "High", "ambiguity_note": "Use sponsor coding convention consistently."
        },
        {
            "spec_seq": 10, "target_domain": "AE", "target_variable": "AESEV",
            "target_label": "Severity/Intensity", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SEVERITY_RAW",
            "source_label": "Raw Severity", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (AESEV: MILD | MODERATE | SEVERE)", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "AE_TOXGR_RAW", "codelist_ref": "AESEV",
            "rule": "Normalize AE_SEVERITY_RAW to approved AESEV values MILD, MODERATE, or SEVERE using deterministic sponsor-controlled terminology mapping only.",
            "primary_qc_checks": "If populated, AESEV must be an approved controlled term and should not clinically contradict AETOXGR when a grade is present.",
            "traceability_note": "Traceable to collected severity field.",
            "programming_notes": "Do not infer severity from toxicity grade when severity is missing; send to exception.",
            "definexml_origin_comment": "Collected severity normalized to controlled terminology.",
            "exception_condition": "Missing, invalid, or clinically contradictory severity -> exception/review",
            "final_condition": "Blank or approved severity term", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 11, "target_domain": "AE", "target_variable": "AETOXGR",
            "target_label": "Toxicity Grade", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 4,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_TOXGR_RAW",
            "source_label": "Raw Toxicity Grade", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Sponsor-controlled toxicity grade codelist or CTCAE-aligned study grading scale", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "AE_SEVERITY_RAW", "codelist_ref": "AETOXGR",
            "rule": "Normalize AE_TOXGR_RAW to sponsor-approved toxicity grade text. In this demo, accepted grades are character values 1-5.",
            "primary_qc_checks": "If populated, AETOXGR must be approved and should not clinically contradict AESEV according to sponsor-defined grade/severity consistency rules.",
            "traceability_note": "Traceable to collected toxicity grade field.",
            "programming_notes": "Treat toxicity grade as character unless sponsor-standardized numeric handling is explicitly required.",
            "definexml_origin_comment": "Collected toxicity grade normalized to sponsor-approved representation.",
            "exception_condition": "Invalid or contradictory toxicity grade -> exception/review",
            "final_condition": "Blank or approved grade", "review_status": "Final", "confidence": "High", "ambiguity_note": "CTCAE-grade semantics may be indication-specific."
        },
        {
            "spec_seq": 12, "target_domain": "AE", "target_variable": "AESER",
            "target_label": "Serious Event", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_RAW",
            "source_label": "Raw Seriousness Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "AE_SER_DTH_RAW + AE_SER_LIFE_RAW + AE_SER_HOSP_RAW + AE_SER_DISAB_RAW + AE_SER_CONG_RAW + AE_SER_MIE_RAW", "codelist_ref": "NY",
            "rule": "Map AESER = Y when AE_SER_RAW indicates a serious adverse event and AESER = N when the event is non-serious, after sponsor-controlled yes/no normalization.",
            "primary_qc_checks": "AESER must align with collected seriousness subcriteria when those are captured. Missing or contradictory seriousness should route to exception.",
            "traceability_note": "Traceable to overall serious/non-serious field and seriousness detail flags.",
            "programming_notes": "Do not derive AESER solely from a sub-criterion when the overall seriousness field is missing; route to exception unless sponsor policy explicitly allows derivation.",
            "definexml_origin_comment": "Collected seriousness normalized to NY codelist.",
            "exception_condition": "Missing, invalid, or contradictory seriousness -> exception",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 13, "target_domain": "AE", "target_variable": "AESDTH",
            "target_label": "Results in Death", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_DTH_RAW",
            "source_label": "Seriousness Criterion: Death", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER + AE_OUTCOME_RAW", "codelist_ref": "NY",
            "rule": "Map AESDTH from AE_SER_DTH_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESDTH=Y, AESER should generally be Y and AEOUT may be FATAL depending on sponsor process.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Do not auto-derive from AEOUT alone.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness or outcome -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 14, "target_domain": "AE", "target_variable": "AESLIFE",
            "target_label": "Is Life Threatening", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_LIFE_RAW",
            "source_label": "Seriousness Criterion: Life Threatening", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER", "codelist_ref": "NY",
            "rule": "Map AESLIFE from AE_SER_LIFE_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESLIFE=Y, AESER should generally be Y.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Do not infer from severity alone.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 15, "target_domain": "AE", "target_variable": "AESHOSP",
            "target_label": "Requires or Prolongs Hospitalization", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_HOSP_RAW",
            "source_label": "Seriousness Criterion: Hospitalization", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER", "codelist_ref": "NY",
            "rule": "Map AESHOSP from AE_SER_HOSP_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESHOSP=Y, AESER should generally be Y.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Do not infer from comments only.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 16, "target_domain": "AE", "target_variable": "AESDISAB",
            "target_label": "Persist or Signif Disability/Incapacity", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_DISAB_RAW",
            "source_label": "Seriousness Criterion: Disability", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER", "codelist_ref": "NY",
            "rule": "Map AESDISAB from AE_SER_DISAB_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESDISAB=Y, AESER should generally be Y.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Do not infer from severity alone.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 17, "target_domain": "AE", "target_variable": "AESCONG",
            "target_label": "Congenital Anomaly or Birth Defect", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_CONG_RAW",
            "source_label": "Seriousness Criterion: Congenital Anomaly", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER", "codelist_ref": "NY",
            "rule": "Map AESCONG from AE_SER_CONG_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESCONG=Y, AESER should generally be Y.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Leave blank when not collected or not applicable.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 18, "target_domain": "AE", "target_variable": "AESMIE",
            "target_label": "Other Medically Important Serious Event", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SER_MIE_RAW",
            "source_label": "Seriousness Criterion: Other Medically Important Event", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "Populate when this seriousness detail is collected", "method_id": "",
            "reconciliation_source": "AESER", "codelist_ref": "NY",
            "rule": "Map AESMIE from AE_SER_MIE_RAW after yes/no normalization.",
            "primary_qc_checks": "If AESMIE=Y, AESER should generally be Y.",
            "traceability_note": "Direct trace to collected seriousness detail flag.",
            "programming_notes": "Leave blank when not collected or not applicable.",
            "definexml_origin_comment": "Collected seriousness detail normalized to NY codelist.",
            "exception_condition": "Contradiction with overall seriousness -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 19, "target_domain": "AE", "target_variable": "AEREL",
            "target_label": "Causality", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_REL_STUDY_DRUG_RAW",
            "source_label": "Relationship to Primary Study Treatment", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Sponsor-approved AEREL codelist aligned to submission terminology", "value_level_condition": "Primary study treatment only in main AE", "method_id": "",
            "reconciliation_source": "AE_REL_STUDY_DRUG2_RAW + RELREC/SUPPAE strategy", "codelist_ref": "AE_REL",
            "rule": "Normalize AE_REL_STUDY_DRUG_RAW to the sponsor-approved AEREL representation for the primary study treatment. Secondary treatment relationship, if collected, should be retained in SUPPAE or a sponsor-defined related-record strategy rather than silently collapsed into AEREL.",
            "primary_qc_checks": "If populated, AEREL must map from an approved relationship term. Multi-treatment relationships require explicit sponsor handling.",
            "traceability_note": "Traceable to primary relationship field; secondary relationship handled separately.",
            "programming_notes": "Do not guess a combined relationship from multiple drug relationship fields.",
            "definexml_origin_comment": "Collected causality normalized to sponsor terminology.",
            "exception_condition": "Unsupported relationship term or unresolved multi-treatment conflict -> exception/review",
            "final_condition": "Blank or approved primary-treatment causality", "review_status": "Final", "confidence": "High", "ambiguity_note": "If the study has more than one investigational treatment, define a sponsor-approved relationship strategy."
        },
        {
            "spec_seq": 20, "target_domain": "AE", "target_variable": "AEACN",
            "target_label": "Action Taken with Study Treatment", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_ACTION_DRUG_RAW",
            "source_label": "Action Taken with Primary Study Treatment", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Sponsor-approved AEACN codelist aligned to submission terminology", "value_level_condition": "Primary study treatment only in main AE", "method_id": "",
            "reconciliation_source": "AE_ACTION_DRUG2_RAW + AE_ACTION_OTHER_TXT + RELREC/SUPPAE strategy", "codelist_ref": "AEACN",
            "rule": "Normalize AE_ACTION_DRUG_RAW to sponsor-approved AEACN values for the primary study treatment. Secondary treatment action is retained separately in SUPPAE or related-record logic when applicable.",
            "primary_qc_checks": "If populated, AEACN must map from an approved action term and should not contradict AEREL or ongoing/outcome context in a clinically impossible way.",
            "traceability_note": "Traceable to primary study-treatment action field.",
            "programming_notes": "Do not collapse multi-treatment action fields without sponsor-approved prioritization.",
            "definexml_origin_comment": "Collected action taken normalized to controlled terminology.",
            "exception_condition": "Unsupported action term or unresolved multi-treatment conflict -> exception/review",
            "final_condition": "Blank or approved primary-treatment action", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 21, "target_domain": "AE", "target_variable": "AEACNOTH",
            "target_label": "Action Taken with Study Treatment, Other", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 200,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_ACTION_OTHER_TXT",
            "source_label": "Free-Text Other Action", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "Populate only when sponsor allows free-text other action in AEACNOTH", "method_id": "",
            "reconciliation_source": "AE_ACTION_DRUG_RAW + AE_ACTION_DRUG2_RAW", "codelist_ref": "",
            "rule": "If the sponsor-approved action mapping uses an OTHER category and AE_ACTION_OTHER_TXT is populated, map the trimmed text to AEACNOTH. Otherwise keep the raw text for operational review or SUPPAE planning only.",
            "primary_qc_checks": "AEACNOTH should not populate unless the corresponding action context supports an 'other' value.",
            "traceability_note": "Direct trace to free-text action field.",
            "programming_notes": "Avoid populating AEACNOTH as a generic comment dump.",
            "definexml_origin_comment": "Collected free-text qualifier when sponsor-approved other-action handling is used.",
            "exception_condition": "Other-text present without valid action context -> exception/review",
            "final_condition": "Blank or sponsor-approved free-text other action", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Sponsor action mapping should explicitly define whether AEACNOTH is populated."
        },
        {
            "spec_seq": 22, "target_domain": "AE", "target_variable": "AEOUT",
            "target_label": "Outcome of Adverse Event", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_OUTCOME_RAW",
            "source_label": "Raw Outcome", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Sponsor-approved AEOUT codelist aligned to submission terminology", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "AE_ONGOING_RAW + AE_SER_DTH_RAW", "codelist_ref": "AEOUT",
            "rule": "Normalize AE_OUTCOME_RAW to the sponsor-approved AEOUT representation using deterministic mapping only.",
            "primary_qc_checks": "Outcome must not contradict ongoing status or fatality seriousness in a clinically impossible way.",
            "traceability_note": "Traceable to collected outcome field.",
            "programming_notes": "Do not infer final outcome solely from end date presence.",
            "definexml_origin_comment": "Collected outcome normalized to controlled terminology.",
            "exception_condition": "Invalid or contradictory outcome -> exception/review",
            "final_condition": "Blank or approved outcome term", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 23, "target_domain": "AE", "target_variable": "AEPRESP",
            "target_label": "Pre-specified Adverse Event", "sdtm_role": "Record Qualifier", "target_type": "Char", "target_length": 1,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_PRESPEC_RAW",
            "source_label": "Pre-specified AE Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (NY)", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "", "codelist_ref": "NY",
            "rule": "Map AEPRESP from AE_PRESPEC_RAW after yes/no normalization.",
            "primary_qc_checks": "If populated, AEPRESP must map to Y/N.",
            "traceability_note": "Direct trace to collected pre-specified AE flag.",
            "programming_notes": "Leave blank when not collected or not applicable.",
            "definexml_origin_comment": "Collected indicator normalized to NY codelist.",
            "exception_condition": "Invalid pre-specified flag -> exception/review",
            "final_condition": "Blank or Y/N", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 24, "target_domain": "AE", "target_variable": "AESTDTC",
            "target_label": "Start Date/Time of Adverse Event", "sdtm_role": "Timing", "target_type": "Char", "target_length": 25,
            "core": "Req", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_START_DATE_RAW + AE_START_TIME_RAW",
            "source_label": "Raw Start Date and Time", "source_type_hint": "date/time-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "ISO 8601 datetime or partial date", "value_level_condition": "", "method_id": "AE_AESTDTC_01",
            "reconciliation_source": "VISITDT_RAW + AE_REPORT_DATE_RAW", "codelist_ref": "",
            "rule": "If AE_START_DATE_RAW is valid ISO 8601 partial date or date, preserve its collected granularity. If a valid start time is also present and sponsor policy allows date-time output, append the normalized time component. If the raw date is non-ISO but valid and deterministic, standardize to ISO 8601 while preserving known partial-date granularity. Invalid or impossible dates route to exception.",
            "primary_qc_checks": "AESTDTC must not be missing for final AE records. Start date/time must not occur after end date/time when both are clean and comparable.",
            "traceability_note": "Traceable to collected start date and time fields.",
            "programming_notes": "Preserve partial dates when only year or year-month is known; do not invent missing precision. No date imputation occurs in the SDTM build from this spec.",
            "definexml_origin_comment": "Collected date/time standardized to ISO 8601 with retained precision.",
            "exception_condition": "Missing, invalid, or contradictory start timing -> exception",
            "final_condition": "Valid ISO 8601 date, partial date, or datetime", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 25, "target_domain": "AE", "target_variable": "AEENDTC",
            "target_label": "End Date/Time of Adverse Event", "sdtm_role": "Timing", "target_type": "Char", "target_length": 25,
            "core": "Perm", "origin": "Collected", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_END_DATE_RAW + AE_END_TIME_RAW",
            "source_label": "Raw End Date and Time", "source_type_hint": "date/time-like char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "ISO 8601 datetime or partial date", "value_level_condition": "Populate when the event is not ongoing and end timing is available", "method_id": "AE_AEENDTC_01",
            "reconciliation_source": "AE_ONGOING_RAW + AE_OUTCOME_RAW", "codelist_ref": "",
            "rule": "If AE_END_DATE_RAW is present and valid, standardize to ISO 8601 while preserving collected precision. Append normalized end time when valid and sponsor policy supports date-time output. If ongoing status indicates the event is ongoing, AEENDTC should generally remain blank.",
            "primary_qc_checks": "AEENDTC should be blank for ongoing events and should not occur before AESTDTC when both are comparable.",
            "traceability_note": "Traceable to collected end date and time fields.",
            "programming_notes": "Do not create an end date from visit date or report date when none was collected. No imputation occurs in the SDTM build from this spec.",
            "definexml_origin_comment": "Collected end date/time standardized to ISO 8601 with retained precision.",
            "exception_condition": "Invalid or contradictory end timing -> exception",
            "final_condition": "Blank or valid ISO 8601 date, partial date, or datetime", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 26, "target_domain": "AE", "target_variable": "AEENRF",
            "target_label": "End Relative to Reference Period", "sdtm_role": "Timing", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_ONGOING_RAW",
            "source_label": "Raw Ongoing Flag", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "CDISC Controlled Terminology (ONGOING)", "value_level_condition": "Populate only when the event is ongoing at cutoff", "method_id": "AE_AEENRF_01",
            "reconciliation_source": "AE_END_DATE_RAW + AE_END_TIME_RAW + AE_OUTCOME_RAW", "codelist_ref": "AEENRF",
            "rule": "If AE_ONGOING_RAW indicates YES and no valid AEENDTC is present, set AEENRF = 'ONGOING'. Otherwise leave AEENRF blank.",
            "primary_qc_checks": "AEENRF should not populate when AEENDTC is populated unless sponsor has a documented exception rule.",
            "traceability_note": "Derived from ongoing status and absence of end timing.",
            "programming_notes": "Use AEENRF rather than manufacturing an end date placeholder.",
            "definexml_origin_comment": "Derived from collected ongoing status.",
            "exception_condition": "Ongoing flag contradicts end date/outcome -> exception/review",
            "final_condition": "Blank or ONGOING", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 27, "target_domain": "AE", "target_variable": "VISIT",
            "target_label": "Visit Name", "sdtm_role": "Timing Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "VISIT_RAW",
            "source_label": "Visit Name Raw", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "Sponsor visit codelist", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "VISITDT_RAW", "codelist_ref": "AE_VISIT",
            "rule": "Map VISIT = trimmed VISIT_RAW preserving the sponsor-defined visit label used for collection and traceability.",
            "primary_qc_checks": "If populated, visit labels should align with the sponsor visit schedule and visit-date context.",
            "traceability_note": "Direct trace to raw visit field.",
            "programming_notes": "Do not silently recode contradictory visit labels.",
            "definexml_origin_comment": "Collected visit label.",
            "exception_condition": "Contradictory or unsupported visit label -> exception/review",
            "final_condition": "Blank or approved visit label", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 28, "target_domain": "AE", "target_variable": "VISITNUM",
            "target_label": "Visit Number", "sdtm_role": "Timing Qualifier", "target_type": "Num", "target_length": 8,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Recode",
            "source_form_or_module": "Adverse Events CRF / Visit map", "source_variable": "VISIT_RAW",
            "source_label": "Visit Name to Visit Number", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "ae_visit_map_v2.csv", "value_level_condition": "", "method_id": "AE_VISITNUM_01",
            "reconciliation_source": "VISITDT_RAW", "codelist_ref": "",
            "rule": "Assign VISITNUM using the sponsor-approved visit map. For this demo, visit numbering is controlled by ae_visit_map_v2.csv rather than a directly collected numeric field.",
            "primary_qc_checks": "If populated, VISITNUM should align one-to-one with the approved visit map for the collected VISIT label.",
            "traceability_note": "Traceable to collected visit label via sponsor-approved visit map.",
            "programming_notes": "Keep unscheduled visit numbering explicit in the visit map; do not derive ad hoc decimals unless sponsor approves them.",
            "definexml_origin_comment": "Assigned via sponsor visit-mapping algorithm.",
            "exception_condition": "Visit label not found in approved visit map -> exception",
            "final_condition": "Blank or approved mapped visit number", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 29, "target_domain": "AE", "target_variable": "EPOCH",
            "target_label": "Epoch", "sdtm_role": "Timing Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Derived", "mapping_class": "Derived",
            "source_form_or_module": "SE or TA domain alignment", "source_variable": "AESTDTC + subject element timing",
            "source_label": "Epoch derivation input", "source_type_hint": "derived", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "Study epochs", "value_level_condition": "Populate only when SE/TA-based derivation is in scope", "method_id": "AE_EPOCH_01",
            "reconciliation_source": "DM/RFSTDTC + SE domain", "codelist_ref": "EPOCH",
            "rule": "Leave EPOCH blank in the standalone AE demo unless the sponsor also provides SE/TA-based epoch boundaries. When in scope, derive EPOCH from AESTDTC according to approved epoch derivation rules.",
            "primary_qc_checks": "If populated, EPOCH should be derived from the start of the event, not the end.",
            "traceability_note": "Derived from subject element timing if available.",
            "programming_notes": "Do not infer EPOCH from visit labels alone.",
            "definexml_origin_comment": "Derived from subject element timing when that dependency is provided.",
            "exception_condition": "Missing SE/TA dependency for intended EPOCH derivation -> leave blank or exception per sponsor policy",
            "final_condition": "Blank or approved epoch value", "review_status": "Final", "confidence": "Medium", "ambiguity_note": "Requires cross-domain dependency outside standalone AE raw."
        },
        {
            "spec_seq": 30, "target_domain": "AE", "target_variable": "AESPID",
            "target_label": "Sponsor-Defined Identifier", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "ROW_ID",
            "source_label": "Raw AE Row Identifier", "source_type_hint": "char", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "", "method_id": "",
            "reconciliation_source": "AE_FORM_SEQ + AE_SEQ_CRf", "codelist_ref": "",
            "rule": "Map AESPID = ROW_ID when the raw row identifier is stable and sponsor wants direct operational traceability from submitted AE to the collected source row.",
            "primary_qc_checks": "If populated, AESPID should be unique enough to support traceability and not change across rebuilds for the same raw snapshot.",
            "traceability_note": "Direct trace to collected operational row identifier.",
            "programming_notes": "If ROW_ID is not stable across source extracts, replace with a sponsor-approved stable operational identifier.",
            "definexml_origin_comment": "Collected sponsor-defined identifier used for traceability.",
            "exception_condition": "Unstable or duplicate ROW_ID where traceability is required -> exception/review",
            "final_condition": "Blank or stable row identifier", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 31, "target_domain": "AE", "target_variable": "AEGRPID",
            "target_label": "Group ID", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 20,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_FORM_SEQ",
            "source_label": "AE Form Sequence", "source_type_hint": "char/numeric-like", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "Populate only if sponsor uses CRF form-sequence as event grouping identifier", "method_id": "",
            "reconciliation_source": "ROW_ID + AE_SEQ_CRf", "codelist_ref": "",
            "rule": "Optionally map AEGRPID = AE_FORM_SEQ when sponsor wants to preserve a form/page-level grouping identifier across related event rows. Otherwise leave blank.",
            "primary_qc_checks": "If populated, AEGRPID should be stable within the intended grouping context and not be confused with AESEQ.",
            "traceability_note": "Traceable to collected AE form sequence.",
            "programming_notes": "Useful only if sponsor's collection design uses an operational grouping concept.",
            "definexml_origin_comment": "Collected sponsor grouping identifier when retained.",
            "exception_condition": "Unclear or unstable grouping use case -> leave blank",
            "final_condition": "Blank or stable group ID", "review_status": "Final", "confidence": "Medium", "ambiguity_note": ""
        },
        {
            "spec_seq": 32, "target_domain": "AE", "target_variable": "AEREFID",
            "target_label": "Reference ID", "sdtm_role": "Identifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Collected", "mapping_class": "Direct",
            "source_form_or_module": "Adverse Events CRF", "source_variable": "AE_SEQ_CRf",
            "source_label": "CRF AE Sequence", "source_type_hint": "char/numeric-like", "source_role_in_rule": "Primary",
            "controlled_terms_or_format": "", "value_level_condition": "Populate only if sponsor wants to retain the raw CRF event line identifier", "method_id": "",
            "reconciliation_source": "ROW_ID + AE_FORM_SEQ", "codelist_ref": "",
            "rule": "Optionally map AEREFID = AE_SEQ_CRf or a formatted representation of it when the sponsor wants to preserve the collected event-line identifier for traceability.",
            "primary_qc_checks": "If populated, AEREFID should not be assumed to be unique in final AE without review; raw duplication should be logged by QC.",
            "traceability_note": "Traceable to raw AE sequence/line identifier.",
            "programming_notes": "Keep as traceability only; do not use as final AESEQ.",
            "definexml_origin_comment": "Collected operational reference identifier when retained.",
            "exception_condition": "Unstable or conflicting raw line identifier -> blank or review",
            "final_condition": "Blank or traceability reference id", "review_status": "Final", "confidence": "High", "ambiguity_note": ""
        },
        {
            "spec_seq": 33, "target_domain": "AE", "target_variable": "AEDICTV",
            "target_label": "Dictionary Version", "sdtm_role": "Variable Qualifier", "target_type": "Char", "target_length": 40,
            "core": "Perm", "origin": "Assigned", "mapping_class": "Constant",
            "source_form_or_module": "Coding workflow metadata", "source_variable": "MedDRA dictionary version",
            "source_label": "Dictionary version metadata", "source_type_hint": "char", "source_role_in_rule": "Support",
            "controlled_terms_or_format": "", "value_level_condition": "Populate only when sponsor carries dictionary version at record level; otherwise document at define.xml level only", "method_id": "",
            "reconciliation_source": "MedDRA coding output", "codelist_ref": "",
            "rule": "If sponsor policy requires record-level dictionary version, populate AEDICTV from the approved MedDRA version metadata. Otherwise leave blank and capture dictionary version in define.xml external codelist metadata.",
            "primary_qc_checks": "If populated, AEDICTV should match the version used for AEDECOD/AESOC coding.",
            "traceability_note": "Traceable to coding metadata, not raw CRF.",
            "programming_notes": "Many sponsors document dictionary version at the metadata level instead of row level.",
            "definexml_origin_comment": "Assigned from coding metadata when retained at record level.",
            "exception_condition": "Coded terms without consistent dictionary version metadata -> coding review",
            "final_condition": "Blank or approved dictionary version", "review_status": "Final", "confidence": "Medium", "ambiguity_note": ""
        }
    ]
    return pd.DataFrame(rows)


def build_visit_map():
    return pd.DataFrame([
        ["SCREENING", 10, "N", "Typical pre-treatment visit"],
        ["BASELINE", 20, "N", "Baseline visit if collected in AE context"],
        ["DAY 1", 20, "N", "Treatment start / baseline-adjacent visit depending on protocol"],
        ["WEEK 1", 30, "N", ""],
        ["WEEK 2", 40, "N", ""],
        ["WEEK 4", 50, "N", ""],
        ["WEEK 8", 60, "N", ""],
        ["UNSCHEDULED", 99, "Y", "Use sponsor-approved unscheduled convention"],
    ], columns=["visit_name", "visitnum", "unsched_expected", "note"])


def build_relationship_action_map():
    return pd.DataFrame([
        ["AE_REL_STUDY_DRUG_RAW", "NOT RELATED", "NOT RELATED", "Primary study treatment relationship"],
        ["AE_REL_STUDY_DRUG_RAW", "POSSIBLY RELATED", "POSSIBLY RELATED", "Primary study treatment relationship"],
        ["AE_REL_STUDY_DRUG_RAW", "RELATED", "RELATED", "Primary study treatment relationship"],
        ["AE_ACTION_DRUG_RAW", "NONE", "DOSE NOT CHANGED", "Map sponsor term NONE to standard action meaning no dose/action change"],
        ["AE_ACTION_DRUG_RAW", "DOSE NOT CHANGED", "DOSE NOT CHANGED", ""],
        ["AE_ACTION_DRUG_RAW", "DOSE REDUCED", "DOSE REDUCED", ""],
        ["AE_ACTION_DRUG_RAW", "DRUG INTERRUPTED", "DRUG INTERRUPTED", ""],
        ["AE_ACTION_DRUG_RAW", "DRUG WITHDRAWN", "DRUG WITHDRAWN", ""],
        ["AE_OUTCOME_RAW", "RECOVERED/RESOLVED", "RECOVERED/RESOLVED", ""],
        ["AE_OUTCOME_RAW", "RECOVERING/RESOLVING", "RECOVERING/RESOLVING", ""],
        ["AE_OUTCOME_RAW", "NOT RECOVERED/NOT RESOLVED", "NOT RECOVERED/NOT RESOLVED", ""],
        ["AE_OUTCOME_RAW", "FATAL", "FATAL", ""],
        ["AE_OUTCOME_RAW", "UNKNOWN", "UNKNOWN", ""],
    ], columns=["source_variable", "raw_value", "target_value", "mapping_note"])


def build_codelists():
    return pd.DataFrame([
        ["DOMAIN", "Extensible codelist", "AE"],
        ["NY", "CDISC codelist", "Y | N"],
        ["AESEV", "CDISC codelist", "MILD | MODERATE | SEVERE"],
        ["AETOXGR", "Sponsor codelist", "1 | 2 | 3 | 4 | 5"],
        ["AE_REL", "Sponsor codelist", "NOT RELATED | POSSIBLY RELATED | RELATED"],
        ["AEACN", "Sponsor codelist", "DOSE NOT CHANGED | DOSE REDUCED | DRUG INTERRUPTED | DRUG WITHDRAWN"],
        ["AEOUT", "Sponsor codelist", "RECOVERED/RESOLVED | RECOVERING/RESOLVING | NOT RECOVERED/NOT RESOLVED | FATAL | UNKNOWN"],
        ["AEPRESP", "CDISC codelist", "Y | N"],
        ["VISIT", "Sponsor codelist", "Controlled per ae_visit_map_v2.csv"],
        ["AEDECOD", "External codelist", "MedDRA Preferred Term"],
        ["AESOC", "External codelist", "MedDRA SOC"],
        ["AEBODSYS", "External codelist", "MedDRA SOC"],
        ["AEENRF", "CDISC codelist", "ONGOING"],
    ], columns=["variable", "codelist_type", "permitted_values_or_reference"])


def build_value_level_metadata():
    return pd.DataFrame([
        ["AEMODIFY", "Populate only when reported term is intentionally modified to facilitate coding", "Char", "Coding-review dependent variable", "Coding governance"],
        ["AEDECOD", "Populate only when MedDRA coding completed; mandatory in final submission AE dataset", "Char", "Preferred term from coding output", "MedDRA coding output"],
        ["AESOC", "Populate only when MedDRA coding completed", "Char", "Primary SOC from coding output", "MedDRA coding output"],
        ["AEBODSYS", "Populate only when MedDRA coding completed", "Char", "Body system/SOC from coding output", "MedDRA coding output"],
        ["AEENDTC", "Populate when event is not ongoing and end timing is available", "Char", "Preserve collected end-date precision", "AE_END_DATE_RAW + AE_END_TIME_RAW"],
        ["AEENRF", "Populate when AE_ONGOING_RAW=YES and no valid AEENDTC is present", "Char", "Use value ONGOING only", "AE_ONGOING_RAW"],
        ["AEACNOTH", "Populate only when sponsor-approved action mapping uses an OTHER category and free text is present", "Char", "Not every study will use this variable", "AE_ACTION_OTHER_TXT + sponsor action policy"],
        ["EPOCH", "Populate only when SE/TA domain timing is available", "Char", "Requires cross-domain derivation", "SE/TA dependencies"],
        ["AEGRPID", "Populate only when sponsor wants operational grouping carried forward", "Char", "Optional traceability variable", "AE_FORM_SEQ"],
        ["AEREFID", "Populate only when sponsor wants raw line identifier traceability", "Char", "Optional traceability variable", "AE_SEQ_CRf"],
        ["AEDICTV", "Populate only when sponsor retains dictionary version at record level", "Char", "Otherwise capture in define.xml only", "Coding metadata"],
    ], columns=["variable", "where_clause", "datatype", "comment", "source_or_rule"])


def build_methods():
    return pd.DataFrame([
        ["AE_USUBJID_01", "Derive USUBJID from STUDYID_RAW, SITEID_RAW, and SUBJECT_RAW while preserving character formatting and left padding."],
        ["AE_AESEQ_01", "Assign AESEQ sequentially within USUBJID after final AE record selection, duplicate resolution, coding completion, and collapse-or-retain decision."],
        ["AE_AEMODIFY_01", "Populate AEMODIFY only when sponsor-approved coding workflow intentionally modifies the verbatim term to support MedDRA coding."],
        ["AE_MEDDRA_01", "Map AETERM/AEMODIFY to MedDRA preferred term and SOC fields using validated external coding output. No guessing permitted in transformation code."],
        ["AE_AESTDTC_01", "Convert raw start date/time to ISO 8601 while preserving collected precision and partial-date granularity. No imputation."],
        ["AE_AEENDTC_01", "Convert raw end date/time to ISO 8601 while preserving collected precision and partial-date granularity. No imputation."],
        ["AE_AEENRF_01", "Derive AEENRF='ONGOING' when ongoing flag is yes and no valid end date/time is present."],
        ["AE_VISITNUM_01", "Map VISIT to VISITNUM using sponsor-approved visit map; do not derive unsupported ad hoc visit numbers."],
        ["AE_EPOCH_01", "Derive EPOCH using SE/TA-based epoch boundaries from AESTDTC when those dependencies are available."],
    ], columns=["method_id", "description"])


def build_coding_dependency_plan():
    return pd.DataFrame([
        ["MedDRA coding input", "AETERM and, if applicable, AEMODIFY", "Coding vendor or internal coding team", "Required before final submission AE dataset is generated. No uncoded AE records allowed in final submission datasets."],
        ["Dictionary version", "MedDRA version metadata", "Coding governance", "Capture in define.xml external codelist metadata even if not carried at row level."],
        ["Recoding review", "Updated terms or coding changes after initial lock", "Safety / data management / coding team", "Changes should be auditable and should trigger downstream AE refresh as needed."],
        ["Uncoded records", "Rows with pending or ambiguous coding", "Exception queue", "Do not guess coding in the transformation program. Keep out of final submission output."],
    ], columns=["dependency_topic", "required_input", "owner", "implementation_note"])


def build_support_qc():
    return pd.DataFrame([
        ["AEYN_RAW", "Record eligibility and contradiction handling", "Used to decide whether an AE record should exist at all"],
        ["VISITDT_RAW", "Timing reconciliation", "Supports chronology review against AESTDTC/AEENDTC"],
        ["AE_START_TIME_RAW", "Timing precision", "Supports AESTDTC date-time derivation when valid"],
        ["AE_END_TIME_RAW", "Timing precision", "Supports AEENDTC date-time derivation when valid"],
        ["SCREENING_NO", "Subject traceability", "Candidate for SUPPAE or reviewer traceability only"],
        ["RAND_NO", "Subject traceability", "Candidate for SUPPAE or reviewer traceability only"],
        ["AE_FORM_SEQ", "Operational grouping", "Candidate for AEGRPID or reviewer traceability only"],
        ["AE_SEQ_CRf", "Operational reference id", "Candidate for AEREFID or reviewer traceability only"],
        ["AE_REL_STUDY_DRUG2_RAW", "Secondary treatment relationship", "Not directly mapped to main AE in this demo; use SUPPAE or sponsor related-record strategy"],
        ["AE_ACTION_DRUG2_RAW", "Secondary treatment action", "Not directly mapped to main AE in this demo; use SUPPAE or sponsor related-record strategy"],
        ["AE_REPORTED_BY", "Operational provenance", "Candidate for SUPPAE only if sponsor wants reporter retained"],
        ["AE_REPORT_DATE_RAW", "Operational reporting chronology", "Used for QC/exceptions; typically not mapped to standard AE variable in this demo"],
        ["AE_COMMENT", "Operational review notes", "QC/exceptions and optional SUPPAE planning"],
        ["ENTRY_STATUS_RAW", "Workflow/audit support", "Operational review only"],
        ["CHANGE_REASON_RAW", "Workflow/audit support", "Operational review only"],
    ], columns=["source_variable", "purpose", "recommended_handling"])


def build_supplemental_qualifier_plan():
    return pd.DataFrame([
        ["SCREENING_NO", "SUPPAE", "AESCNNO", "Screening number", "Candidate only if sponsor wants retained operational screening identifier"],
        ["RAND_NO", "SUPPAE", "AERANDNO", "Randomization number", "Candidate only if sponsor wants retained randomization identifier in AE context"],
        ["AE_FORM_SEQ", "SUPPAE", "AEFORMSQ", "AE form sequence", "Useful for source-row traceability when AEGRPID is not used"],
        ["AE_SEQ_CRf", "SUPPAE", "AECRFSEQ", "CRF AE sequence", "Useful for source traceability when AEREFID is not used"],
        ["AE_REL_STUDY_DRUG2_RAW", "SUPPAE", "AEREL2", "Relationship to secondary study treatment", "Use only if sponsor retains multi-treatment causality details in SUPPAE"],
        ["AE_ACTION_DRUG2_RAW", "SUPPAE", "AEACN2", "Action taken with secondary study treatment", "Use only if sponsor retains multi-treatment action details in SUPPAE"],
        ["AE_REPORTED_BY", "SUPPAE", "AEREPOR", "Reported by", "Operational provenance only if sponsor wants retained reporter role"],
        ["AE_REPORT_DATE_RAW", "SUPPAE", "AERPTDT", "AE report date", "Operational reporting chronology only if sponsor wants retained report date"],
        ["ENTRY_STATUS_RAW", "SUPPAE", "AEENTSTS", "Entry status", "Usually reviewer traceability only, not submission-essential"],
        ["CHANGE_REASON_RAW", "SUPPAE", "AECHGREAS", "Change reason", "Usually reviewer traceability only, not submission-essential"],
        ["AE_COMMENT", "SUPPAE", "AECOMM", "AE comment", "Only if sponsor explicitly wants free-text operational comments retained"],
    ], columns=["raw_variable", "supp_dataset", "qnam", "qlabel", "note"])


def build_relrec_strategy():
    return pd.DataFrame([
        ["Primary study treatment relationship/action only", "Populate AEREL and AEACN from primary treatment fields", "No RELREC needed for main AE relationship handling"],
        ["Secondary treatment relationship collected", "Retain secondary relationship in SUPPAE or related record strategy", "Do not silently collapse into AEREL"],
        ["Secondary treatment action collected", "Retain secondary action in SUPPAE or related record strategy", "Do not silently collapse into AEACN"],
        ["Study with multiple investigational products and formal exposure linkage", "Implement RELREC between AE and exposure records according to sponsor design", "Main AE should still retain clear primary mapping logic"],
    ], columns=["scenario", "recommended_handling", "note"])


def build_row_selection_rules():
    return pd.DataFrame([
        ["1", "Eligibility gate", "Only rows that pass Layer 1 QC and deterministic mapping requirements are eligible for final AE. Rows with unresolved contradictions remain in exception outputs."],
        ["2", "AE presence logic", "Rows where AEYN_RAW indicates no event should not generate final AE records unless sponsor-approved exception handling explicitly applies."],
        ["3", "Coding dependency gate", "Rows without completed MedDRA coding may remain in a pre-coding exception/interim extract but should not enter the final submission-ready coded AE dataset."],
        ["4", "Operational duplicates", "Potential duplicates are identified using subject, verbatim/coded term, start timing, seriousness, and sponsor traceability identifiers. Do not drop competing rows without documented row-selection rules."],
        ["5", "Same-event collapse option", "If sponsor policy is to submit one record per unique event from start to finish, collapse operational rows to the highest severity, highest seriousness, most related causality, and final outcome while preserving source traceability."],
        ["6", "Separate-record option", "If sponsor policy treats changes in seriousness, severity, or causality as distinct reportable events, retain separate AE rows instead of collapsing."],
        ["7", "Ongoing event handling", "For ongoing events, do not manufacture AEENDTC. Use AEENRF='ONGOING' when the ongoing flag is clean and the row otherwise qualifies."],
        ["8", "Multi-treatment causality/action", "Where relationship/action is collected for more than one study treatment, apply sponsor-approved main-AE versus SUPPAE/RELREC strategy. Do not silently collapse secondary treatment information."],
        ["9", "Traceability identifiers", "AESPID/AEGRPID/AEREFID may be retained for traceability if sponsor policy requires them, but none of these replace AESEQ in the final dataset."],
        ["10", "Sequence assignment point", "Assign AESEQ only after final record selection and final collapse-or-retain decisions are complete."],
    ], columns=["step_no", "topic", "rule"])


def build_exception_routing():
    return pd.DataFrame([
        ["Missing or contradictory AETERM", "Route to AE exception dataset", "Cannot create trustworthy AE event record"],
        ["Missing/invalid start timing", "Route to AE exception dataset", "AESTDTC is required for final AE build"],
        ["Start/end chronology contradiction", "Route to AE exception dataset", "Timing conflict cannot be trusted automatically"],
        ["Missing/invalid seriousness with seriousness-detail conflict", "Route to AE exception dataset", "Clinical contradiction requires human review"],
        ["Invalid severity or toxicity grade mismatch", "Route to AE exception dataset", "Clinical contradiction requires human review"],
        ["Outcome/ongoing contradiction", "Route to AE exception dataset", "Cannot trust final event disposition automatically"],
        ["Uncoded or ambiguous MedDRA mapping", "Route to coding queue / AE exception dataset", "Cannot populate coded AE reliably"],
        ["Duplicate or competing source records", "Route to AE exception dataset", "Requires explicit sponsor row-selection decision"],
        ["Unresolved multi-treatment relationship/action handling", "Route to AE exception dataset", "Main AE vs SUPPAE/RELREC strategy not deterministic"],
    ], columns=["issue_pattern", "routing_action", "reason"])


def build_qc_rule_crosswalk(rules_cfg):
    if not rules_cfg:
        return pd.DataFrame(columns=["rule_id","variable_focus","severity","final_bucket","rule_description","classification_basis"])
    rows = []
    variable_map = {
        "AE001": "Required fields",
        "AE002": "STUDYID_RAW",
        "AE003": "USUBJID components",
        "AE004": "AEYN_RAW",
        "AE005": "AETERM",
        "AE006": "AESTDTC",
        "AE007": "AEENDTC",
        "AE008": "AESTDTC/AEENDTC",
        "AE009": "AESEV",
        "AE010": "AETOXGR",
        "AE011": "AESEV/AETOXGR",
        "AE012": "AESER",
        "AE013": "Seriousness detail flags",
        "AE014": "AEREL",
        "AE015": "AEACN",
        "AE016": "AEOUT",
        "AE017": "AEPRESP",
        "AE018": "VISIT",
        "AE019": "AESTDTC/VISITDT_RAW",
        "AE020": "AE report chronology",
        "AE021": "AEENRF/AEENDTC",
        "AE022": "AESTDTC/AEENDTC time chronology",
        "AE023": "Potential duplicate AE record",
        "AE024": "Secondary relationship/action",
        "AE035": "AEREFID / AESEQ traceability"
    }
    for rule_id, meta in sorted(rules_cfg.get("rules", {}).items()):
        rows.append({
            "rule_id": rule_id,
            "variable_focus": variable_map.get(rule_id, ""),
            "severity": meta.get("severity"),
            "final_bucket": meta.get("bucket"),
            "rule_description": meta.get("description"),
            "classification_basis": meta.get("basis")
        })
    return pd.DataFrame(rows)


def build_source_profile(source_df):
    if source_df is None:
        return pd.DataFrame(columns=["source_variable","present_in_file","non_null_count","distinct_non_null_count","sample_values"])
    rows = []
    for col in EXPECTED_RAW_COLUMNS:
        present = col in source_df.columns
        if not present:
            rows.append({
                "source_variable": col,
                "present_in_file": "No",
                "non_null_count": None,
                "distinct_non_null_count": None,
                "sample_values": None
            })
            continue
        ser = source_df[col]
        clean = ser.dropna().astype(str).str.strip()
        clean = clean[clean != ""]
        samples = " | ".join(clean.drop_duplicates().head(5).tolist())
        rows.append({
            "source_variable": col,
            "present_in_file": "Yes",
            "non_null_count": int(clean.shape[0]),
            "distinct_non_null_count": int(clean.nunique()),
            "sample_values": samples
        })
    return pd.DataFrame(rows)


def build_definexml_notes():
    return pd.DataFrame([
        ["Dataset metadata", "Dataset label, class, structure, keys, sort order, and clear description of whether AE is collapsed or row-preserving."],
        ["Variable metadata", "Name, label, datatype, length/significant digits, origin, codelist reference, method id, comments."],
        ["Value-level metadata", "Needed for variables with conditional population such as AEMODIFY, AEENDTC, AEENRF, EPOCH, and coding-dependent variables."],
        ["Codelists", "Controlled terminology should be documented for NY, AESEV, toxicity grade, relationship, action, outcome, visit, and MedDRA external codelists."],
        ["Origin", "Use Collected, Derived, Assigned consistently; coding-derived variables should be clearly marked Assigned/Recode."],
        ["Methods", "Derived variables should reference reproducible method identifiers for SDTM generation and define.xml computation methods."],
        ["External links", "Annotated CRF, coding conventions, reviewer guide, and MedDRA dictionary metadata can be attached later outside this builder script."],
    ], columns=["definexml_component", "what_to_capture"])


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 72)

    if ws.max_row >= 2 and ws.max_column >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=f"T_{re.sub(r'[^A-Za-z0-9]', '_', ws.title)}", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)


def write_excel(datasets, out_xlsx):
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"
    readme_rows = [
        ["Sheet", "Purpose"],
        ["Dataset_Metadata", "Dataset-level metadata for AE including class, keys, sort order, coding dependency, and define.xml-ready notes."],
        ["AE_Spec", "Main AE mapping specification; one row per target variable."],
        ["AE_Visit_Map", "Sponsor visit numbering/reference used by AE VISIT/VISITNUM."],
        ["AE_Rel_Action_Map", "Controlled terminology mapping support for relationship, action, and outcome."],
        ["AE_Codelists", "Variable-level codelist planning for define.xml support."],
        ["AE_Value_Level_Metadata", "Conditions where value-level metadata are needed for define.xml."],
        ["Methods", "Method definitions referenced by the main spec and define.xml computation methods."],
        ["AE_Coding_Dependency", "MedDRA coding dependency and governance notes."],
        ["Support_QC", "Operational raw fields used for reconciliation, traceability, and exception routing."],
        ["SUPPAE_Plan", "Candidate supplemental qualifier planning; not mapped directly unless sponsor approves."],
        ["RELREC_Strategy", "Sponsor guidance for multi-treatment causality/action traceability."],
        ["Row_Selection_Rules", "Explicit final record-selection rules for duplicates, collapse strategy, and multi-treatment handling."],
        ["Exception_Routing", "Patterns that route rows to exception outputs rather than final AE."],
        ["QC_Rule_Crosswalk", "Optional linkage between Layer 1 QC rules and the AE spec."],
        ["Source_Profile", "Raw source-column profiling from the detected AE input file."],
        ["DefineXML_Notes", "Checklist for define.xml-ready metadata completion."],
    ]
    for r_idx, row in enumerate(readme_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_readme.cell(r_idx, c_idx).value = value
    style_sheet(ws_readme)

    for name, df in datasets:
        ws = wb.create_sheet(name)
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(1, c_idx).value = col
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(r_idx, c_idx).value = value
        style_sheet(ws)

    wb.save(out_xlsx)


def main():
    OUTDIR.mkdir(exist_ok=True)

    source_path = auto_detect_file(["ae_raw_synthetic.csv", "ae_raw.csv", "ae_source.csv"])
    rules_path = auto_detect_file(["ae_layer1_rules_v6.json", "ae_layer1_rules_v5.json", "ae_layer1_rules_v4.json"])

    source_df = pd.read_csv(source_path, dtype=str) if source_path and source_path.exists() else None
    rules_cfg = read_json(rules_path)

    dataset_metadata = build_dataset_metadata()
    main_spec = build_main_spec()
    visit_map = build_visit_map()
    rel_action_map = build_relationship_action_map()
    codelists = build_codelists()
    vlm = build_value_level_metadata()
    methods = build_methods()
    coding_deps = build_coding_dependency_plan()
    support_qc = build_support_qc()
    suppae_plan = build_supplemental_qualifier_plan()
    relrec_strategy = build_relrec_strategy()
    row_selection = build_row_selection_rules()
    exception_routing = build_exception_routing()
    qc_crosswalk = build_qc_rule_crosswalk(rules_cfg)
    source_profile = build_source_profile(source_df)
    definexml_notes = build_definexml_notes()

    dataset_metadata.to_csv(OUTDIR / "ae_dataset_metadata_v2.csv", index=False)
    main_spec.to_csv(OUTDIR / "ae_mapping_spec_validated_v2.csv", index=False)
    visit_map.to_csv(OUTDIR / "ae_visit_map_v2.csv", index=False)
    rel_action_map.to_csv(OUTDIR / "ae_relationship_action_map_v2.csv", index=False)
    codelists.to_csv(OUTDIR / "ae_codelists_v2.csv", index=False)
    vlm.to_csv(OUTDIR / "ae_value_level_metadata_v2.csv", index=False)
    methods.to_csv(OUTDIR / "ae_methods_v2.csv", index=False)
    coding_deps.to_csv(OUTDIR / "ae_coding_dependency_v2.csv", index=False)
    support_qc.to_csv(OUTDIR / "ae_support_qc_rows_v2.csv", index=False)
    suppae_plan.to_csv(OUTDIR / "ae_suppae_plan_v2.csv", index=False)
    relrec_strategy.to_csv(OUTDIR / "ae_relrec_strategy_v2.csv", index=False)
    row_selection.to_csv(OUTDIR / "ae_row_selection_rules_v2.csv", index=False)
    exception_routing.to_csv(OUTDIR / "ae_exception_routing_v2.csv", index=False)
    qc_crosswalk.to_csv(OUTDIR / "ae_qc_rule_crosswalk_v2.csv", index=False)
    source_profile.to_csv(OUTDIR / "ae_source_profile_v2.csv", index=False)
    definexml_notes.to_csv(OUTDIR / "ae_definexml_notes_v2.csv", index=False)

    write_excel(
        datasets=[
            ("Dataset_Metadata", dataset_metadata),
            ("AE_Spec", main_spec),
            ("AE_Visit_Map", visit_map),
            ("AE_Rel_Action_Map", rel_action_map),
            ("AE_Codelists", codelists),
            ("AE_Value_Level_Metadata", vlm),
            ("Methods", methods),
            ("AE_Coding_Dependency", coding_deps),
            ("Support_QC", support_qc),
            ("SUPPAE_Plan", suppae_plan),
            ("RELREC_Strategy", relrec_strategy),
            ("Row_Selection_Rules", row_selection),
            ("Exception_Routing", exception_routing),
            ("QC_Rule_Crosswalk", qc_crosswalk),
            ("Source_Profile", source_profile),
            ("DefineXML_Notes", definexml_notes),
        ],
        out_xlsx=OUTDIR / "AE_Mapping_Spec_Industry_Grade_v2.xlsx",
    )

    readme = (
        "AE spec package v2\n"
        "- Builds a define.xml-ready AE mapping package aligned to SDTMIG-style AE metadata\n"
        "- Separates coding-dependent variables from raw-to-SDTM transformation logic\n"
        "- Keeps SUPPAE and operational traceability planning separate from the main AE variable spec\n"
        "- Includes row-selection, collapse-vs-separate-record guidance, RELREC strategy, and exception routing\n"
        "- Auto-detects the AE raw source CSV and latest available AE Layer 1 rules JSON when present\n"
        "- Enforces no uncoded AE records in final submission dataset\n"
        "- Ensures controlled terminology compliance and define.xml-ready method metadata\n"
    )
    (OUTDIR / "README.txt").write_text(readme, encoding="utf-8")
    print(f"Created outputs in: {OUTDIR}")
    if source_path:
        print(f"Source CSV used: {source_path.name}")
    if rules_path:
        print(f"Rules JSON used: {rules_path.name}")


if __name__ == "__main__":
    main()
